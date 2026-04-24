#!/usr/bin/env python3
# ═════════════════════════════════════════════════════════════════════
# ⚠️  VERSION DÉPRÉCIÉE — 2026-04-24  ⚠️
# ─────────────────────────────────────────────────────────────────────
# Ce script générait les rapports via ReportLab + matplotlib (2776 lignes,
# layout PDF impératif, graphiques en images). Il a été remplacé par un
# pipeline Jinja2 + Playwright (template HTML A4 éditorial + CSS
# @page Paged Media) : voir ../generate_report.py et ../templates/.
#
# Conservé ici à titre de référence historique — NE PAS UTILISER EN PROD.
# Les rapports Dupont / Lacues / Unia / Dubois générés avec cette version
# sont archivés dans ~/Desktop/ (dossier hors git).
# ═════════════════════════════════════════════════════════════════════
"""
Template de rapport d'estimation de cave à vins
Service d'estimation de cave à vins à distance
"""

import os
import math
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.colors import HexColor, white, black, Color
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Frame, PageTemplate, BaseDocTemplate, KeepTogether,
    Image as RLImage
)
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing, Rect, String, Circle, Wedge, Line as ShapeLine
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics import renderPDF
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np


# ──────────────────────────────────────────────
# COULEURS DU THÈME
# ──────────────────────────────────────────────
WINE_DARK    = HexColor("#2D1B2E")   # Bordeaux très foncé
WINE_PRIMARY = HexColor("#722F37")   # Bordeaux classique
WINE_LIGHT   = HexColor("#A0616D")   # Bordeaux clair
GOLD         = HexColor("#C5A258")   # Or élégant
GOLD_LIGHT   = HexColor("#E8D5A3")   # Or pâle
CREAM        = HexColor("#F2EEE2")   # Fond crème
BG_LIGHT     = HexColor("#F2EEE2")   # Fond légèrement plus soutenu
TEXT_DARK    = HexColor("#1A1A1A")   # Texte principal
TEXT_MID     = HexColor("#4A4A4A")   # Texte secondaire
TEXT_LIGHT   = HexColor("#7A7A7A")   # Texte tertiaire
BORDER_LIGHT = HexColor("#E0D8CC")   # Bordure subtile
GREEN_REC    = HexColor("#2E7D4F")   # Recommandation "garder"
ORANGE_REC   = HexColor("#D4841A")   # Recommandation "attendre"
RED_REC      = HexColor("#B33A3A")   # Recommandation "vendre"

PAGE_W, PAGE_H = A4
MARGIN_L = 16 * mm
MARGIN_R = 16 * mm
MARGIN_T = 20 * mm
MARGIN_B = 25 * mm



# ──────────────────────────────────────────────
# IMPORTS SUPPLÉMENTAIRES
# ──────────────────────────────────────────────
import unicodedata
import re as regex_module
import sys
import csv


def generate_ref_mission(nom):
    """Génère automatiquement la référence mission : EST-YYYY-MM-XXX."""
    cleaned = nom.replace("M.", "").replace("Mme", "").replace("Mme.", "").replace("et", "").strip()
    words = [w for w in cleaned.split() if len(w) > 1]
    last_word = words[-1] if words else "XXX"
    normalized = unicodedata.normalize('NFD', last_word)
    ascii_only = normalized.encode('ascii', 'ignore').decode('ascii')
    letters_only = regex_module.sub(r'[^A-Za-z]', '', ascii_only)
    code = letters_only[:3].upper() if len(letters_only) >= 3 else letters_only.upper().ljust(3, 'X')
    now = datetime.now()
    return f"EST-{now.year}-{now.month:02d}-{code}"


# ──────────────────────────────────────────────
# LECTURE AUTOMATIQUE — CSV TALLY (ÉTAPE 1)
# ──────────────────────────────────────────────
def lire_csv_tally(fichier_csv):
    """Lit le CSV exporté depuis Tally et remplit CLIENT_DATA.
    Correspondances : Nom et prénom → nom, Email → email,
    Objectif de la mission → objectif, Provenance de la cave → provenance,
    Conditions de conservation → conditions_conservation,
    Localisation (ville ou département) → localisation."""
    with open(fichier_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if not rows:
        print("ERREUR : CSV Tally vide")
        sys.exit(1)
    row = rows[-1]  # Dernière soumission = la plus récente

    # Conditions d'estimation — logique dynamique
    mode_inventaire = row.get("Comment souhaitez-vous nous transmettre votre inventaire ?", "").strip().lower()
    has_photos = bool(row.get("Upload de vos photos", "").strip())
    has_drive = bool(row.get("Lien de partage", "").strip())
    suffix_photos = " et photos (étiquettes, niveaux, capsules)" if (has_photos or has_drive) else ""

    if "ligne par ligne" in mode_inventaire or "texte" in mode_inventaire:
        conditions_est = f"Sur description textuelle fournie par le client{suffix_photos}"
    elif "fichier" in mode_inventaire or "upload" in mode_inventaire:
        conditions_est = f"Sur liste et fichier fournis par le client{suffix_photos}"
    elif has_photos or has_drive:
        conditions_est = "Sur photos (étiquettes, niveaux, capsules)"
    else:
        conditions_est = "Sur éléments fournis par le client"

    return {
        "nom": row.get("Nom et prénom", "").strip(),
        "email": row.get("Email", "").strip(),
        "objectif": row.get("Objectif de la mission", "").strip() or "[Non renseigné]",
        "conditions_conservation": row.get("Conditions de conservation", "").strip() or "[Non renseigné]",
        "localisation": row.get("Localisation (ville ou département)", "").strip(),
        "provenance": row.get("Provenance de la cave", "").strip() or "[Non renseigné]",
        "conditions_estimation": conditions_est,
        "infos_complementaires": row.get("En savoir plus", "").strip(),
    }


# ──────────────────────────────────────────────
# LECTURE AUTOMATIQUE — EXCEL INVENTAIRE (ÉTAPE 2)
# ──────────────────────────────────────────────
def lire_inventaire_excel(fichier_excel):
    """Lit un fichier Excel inventaire par nom de colonne.
    Colonnes attendues (dans n'importe quel ordre) :
    Domaine | Appellation | Region | Couleur | Millesime | Format | Qte | Etat |
    CBO | Val_unit | Reco | Justification_code | Duree_garde | Apogee | Note_marche"""
    try:
        import openpyxl
    except ImportError:
        print("ERREUR : openpyxl requis. Installez-le avec : pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(fichier_excel, read_only=True)
    print(f"  Feuilles disponibles : {wb.sheetnames}")
    if "Inventaire" in wb.sheetnames:
        ws = wb["Inventaire"]
        print(f"  → Lecture de la feuille 'Inventaire'")
    else:
        ws = wb.active
        print(f"  → Feuille 'Inventaire' non trouvée, lecture de '{ws.title}'")
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("ERREUR : fichier Excel vide")
        sys.exit(1)

    # Lire les headers par nom — première ligne
    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    print(f"  En-têtes trouvées ({len(raw_headers)}) : {raw_headers}")

    # Mapping flexible : normaliser les en-têtes pour tolérer les variantes
    header_aliases = {
        "domaine": "Domaine", "domaine / cuvée": "Domaine", "domaine/cuvée": "Domaine",
        "domaine / cuvee": "Domaine", "nom": "Domaine", "cuvée": "Domaine",
        "appellation": "Appellation",
        "region": "Region", "région": "Region",
        "couleur": "Couleur", "type": "Couleur",
        "millesime": "Millesime", "millésime": "Millesime", "mill": "Millesime", "mill.": "Millesime",
        "format": "Format",
        "qte": "Qte", "qté": "Qte", "quantite": "Qte", "quantité": "Qte", "nb": "Qte",
        "etat": "Etat", "état": "Etat", "etat apparent": "Etat", "état apparent": "Etat",
        "cbo": "CBO", "caisse bois": "CBO",
        "val_unit": "Val_unit", "val unit": "Val_unit", "valeur unitaire": "Val_unit",
        "val. unit.": "Val_unit", "prix": "Val_unit", "valeur": "Val_unit",
        "reco": "Reco", "recommandation": "Reco", "reco.": "Reco",
        "justification_code": "Justification_code", "justification": "Justification_code",
        "code justification": "Justification_code", "code": "Justification_code",
        "duree_garde": "Duree_garde", "durée de garde": "Duree_garde", "duree garde": "Duree_garde",
        "durée garde": "Duree_garde", "garde": "Duree_garde",
        "apogee": "Apogee", "apogée": "Apogee",
        "note_marche": "Note_marche", "note marché": "Note_marche", "note marche": "Note_marche",
        "note de marché": "Note_marche",
    }

    headers = []
    for h in raw_headers:
        normalized = h.lower().strip()
        mapped = header_aliases.get(normalized, h)
        headers.append(mapped)

    nb_cols = len(headers)
    print(f"  En-têtes normalisées : {headers}")
    inventaire = []

    # Normalisation Reco — variantes sans accent
    reco_map = {"A conserver": "À conserver", "A vendre": "À vendre", "A surveiller": "À surveiller"}

    def safe_str(val, default=""):
        """Convertit une valeur en string nettoyée, None → default."""
        if val is None:
            return default
        s = str(val).strip()
        return default if s.lower() in ("none", "") else s

    for row in rows[1:]:
        if not row:
            continue

        # Padder la ligne si elle a moins de colonnes que les headers
        padded = list(row) + [None] * (nb_cols - len(row))
        d = dict(zip(headers, padded))

        # Ignorer lignes vides, en-têtes en double, et lignes d'instructions
        domaine_val = safe_str(d.get("Domaine"))
        if not domaine_val or domaine_val.lower() == "domaine":
            continue
        # Ignorer les lignes qui ne commencent pas par une lettre ou un chiffre (emoji, symboles)
        if domaine_val and not domaine_val[0].isalnum():
            continue

        # Millésime : NM → 0
        mill_raw = d.get("Millesime")
        if mill_raw is None or str(mill_raw).strip().upper() in ("NM", "", "NONE"):
            mill = 0
        else:
            try:
                mill = int(float(mill_raw))
            except (ValueError, TypeError):
                mill = 0

        # Code justification
        jcode = None
        jcode_raw = d.get("Justification_code")
        if jcode_raw is not None and str(jcode_raw).strip() not in ("", "None"):
            try:
                jcode = int(float(jcode_raw))
            except (ValueError, TypeError):
                jcode = None

        # Apogée — utiliser la valeur Excel si renseignée
        apogee_val = safe_str(d.get("Apogee"))

        # Reco — normaliser
        reco_raw = safe_str(d.get("Reco"), "À surveiller")
        reco_val = reco_map.get(reco_raw, reco_raw)

        # Quantité et valeur unitaire
        try:
            qte = int(float(d.get("Qte") or 1))
        except (ValueError, TypeError):
            qte = 1
        try:
            val_unit = int(float(d.get("Val_unit") or 0))
        except (ValueError, TypeError):
            val_unit = 0

        inventaire.append({
            "bouteille": domaine_val,
            "appellation": safe_str(d.get("Appellation")),
            "region": safe_str(d.get("Region")),
            "couleur": safe_str(d.get("Couleur"), "Rouge"),
            "format": safe_str(d.get("Format"), "75cl"),
            "millesime": mill,
            "apogee": apogee_val,
            "etat": safe_str(d.get("Etat"), "Bon"),
            "cbo": safe_str(d.get("CBO"), "Non"),
            "qte": qte,
            "val_unit": val_unit,
            "reco": reco_val,
            "justification_code": jcode,
            "duree_garde": safe_str(d.get("Duree_garde")) if reco_val == "À conserver" else "",
            "note_marche": safe_str(d.get("Note_marche")),
        })

    print(f"  Lignes de données lues : {len(rows) - 1}, références valides : {len(inventaire)}")
    if len(inventaire) == 0:
        print("  ATTENTION : aucune référence valide trouvée. Vérifiez que :")
        print("    - La feuille 'Inventaire' contient des données à partir de la ligne 2")
        print("    - La colonne 'Domaine' n'est pas vide")
        print(f"    - Les en-têtes correspondent à : {headers}")
    return inventaire


# ══════════════════════════════════════════════════════════════
# VARIABLES D'EXPERTISE — chargées depuis rapport_data.py
# ══════════════════════════════════════════════════════════════
# Placer `rapport_data.py` dans le dossier du client (cwd) avec les 6 variables.
# Si absent, fallback sur le template du skill puis sur des placeholders.

_EXPERT_DEFAULTS = {
    "liquidite_cave": "Moyenne",
    "aide_decision": "[À rédiger — conclusion de l'expert]",
    "points_vigilance": "[À rédiger — points de vigilance]",
    "prochaines_etapes": "[À rédiger — prochaines étapes]",
    "contexte_marche_tendances": "[À rédiger — tendances du marché]",
    "contexte_marche_liquidite": "[À rédiger — liquidité des références]",
}

def _load_expert_vars():
    import importlib.util
    script_dir = os.path.dirname(os.path.abspath(__file__))
    for path in (os.path.join(os.getcwd(), "rapport_data.py"),
                 os.path.join(script_dir, "rapport_data.py")):
        if os.path.exists(path):
            spec = importlib.util.spec_from_file_location("rapport_data", path)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            print(f"  Variables d'expertise chargées depuis : {path}")
            return {k: getattr(mod, k, _EXPERT_DEFAULTS[k]) for k in _EXPERT_DEFAULTS}
    print("  Aucun rapport_data.py trouvé — placeholders utilisés")
    return _EXPERT_DEFAULTS.copy()

_expert = _load_expert_vars()
liquidite_cave = _expert["liquidite_cave"]
aide_decision = _expert["aide_decision"]
points_vigilance = _expert["points_vigilance"]
prochaines_etapes = _expert["prochaines_etapes"]
contexte_marche_tendances = _expert["contexte_marche_tendances"]
contexte_marche_liquidite = _expert["contexte_marche_liquidite"]


# ──────────────────────────────────────────────
# DONNÉES D'EXEMPLE — chargées depuis demo_data.py
# ──────────────────────────────────────────────
# En mode prod (CSV + XLSX fournis), ces valeurs sont écrasées.
def _load_demo_data():
    import importlib.util
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "demo_data.py")
    if os.path.exists(path):
        spec = importlib.util.spec_from_file_location("demo_data", path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod.CLIENT_DATA, mod.INVENTAIRE
    return ({"nom": "Client", "email": "", "objectif": "", "conditions_estimation": "",
             "conditions_conservation": "", "provenance": "", "localisation": ""}, [])

CLIENT_DATA, INVENTAIRE = _load_demo_data()

# ──────────────────────────────────────────────
# DONNÉES FIXES — NE PAS MODIFIER
# ──────────────────────────────────────────────
CLIENT_DATA["date_rapport"] = datetime.now().strftime("%d/%m/%Y")
CLIENT_DATA["ref_mission"] = generate_ref_mission(CLIENT_DATA["nom"])
CLIENT_DATA["contact_email"] = "contact@estimationcave.com"
CLIENT_DATA["type_offre"] = "Audit & Estimation"
CLIENT_DATA["texte_intro_garder"] = "Les lots qu\'il est préférable de garder à moyen ou long terme, parce qu\'ils ont encore du potentiel de garde ou que leur marché n\'est pas encore mature."
CLIENT_DATA["texte_intro_vendre"] = "Les lots qu\'il est pertinent de vendre à court terme, au regard de leur maturité, de leur attractivité sur le marché ou de l\'objectif exprimé dans le cadre de la mission."
CLIENT_DATA["texte_intro_attendre"] = "Les lots pour lesquels il vaut mieux attendre avant de décider, en raison d\'une maturité encore indéterminée ou d\'un manque de visibilité suffisant sur leur attractivité."
CLIENT_DATA["liquidite_cave"] = liquidite_cave
CLIENT_DATA["aide_decision"] = aide_decision
CLIENT_DATA["points_vigilance"] = points_vigilance
CLIENT_DATA["prochaines_etapes"] = prochaines_etapes

# Codes de justification — règles de décision détaillées dans grille_analyse_vin.md
JUSTIFICATION_CODES = {
    # À CONSERVER
    1: "Fenêtre d'apogée non atteinte",
    2: "Cote en progression",
    3: "Référence rare ou confidentielle",
    4: "Millésime exceptionnel",
    5: "Grand format (meilleure conservation)",
    6: "Valeur de dégustation supérieure à la valeur de revente",
    # À VENDRE
    7: "Apogée atteinte ou dépassée",
    8: "Fenêtre de vente favorable",
    9: "État non compatible avec une conservation prolongée",
    10: "Référence accessible, ne se valorisera pas à l'avenir",
    11: "Millésime anniversaire (prime de liquidité)",
    12: "Format demi-bouteille (moins bonne conservation)",
    17: "Cote ayant atteint un plateau / tendance baissière",
    # À SURVEILLER
    13: "Maturité encore indéterminée",
    14: "Attractivité encore insuffisante sur le marché",
    15: "Millésime illisible ou inconnu — décision impossible sans vérification",
    16: "Signaux de marché contradictoires — tendance à confirmer",
}

# Totaux calculés
def total_btls():
    return sum(b["qte"] for b in INVENTAIRE)

def total_val():
    return sum(b["qte"] * b["val_unit"] for b in INVENTAIRE)


def estimate_apogee(b):
    """Estime automatiquement la fenêtre d'apogée d'une bouteille.
    Basé sur la région, l'appellation, le millésime et l'état.
    Retourne 'AAAA–AAAA', 'À boire maintenant' ou 'À évaluer'."""
    mill = b.get("millesime", 0)
    if mill == 0:
        return "À boire maintenant"

    region = b.get("region", "").lower()
    appellation = b.get("appellation", "").lower()
    etat = b.get("etat", "Bon")

    # Durée de garde typique par profil (années après millésime)
    # [début_apogée, fin_apogée]
    if "grand cru" in appellation or "1er cru" in appellation or "premier cru" in appellation:
        window = (12, 35)
    elif region == "bourgogne":
        window = (8, 25)
    elif "pauillac" in appellation or "margaux" in appellation or "saint-julien" in appellation:
        window = (12, 35)
    elif "sauternes" in appellation:
        window = (10, 50)
    elif region == "bordeaux":
        window = (8, 25)
    elif "châteauneuf" in appellation or "hermitage" in appellation or "côte-rôtie" in appellation:
        window = (8, 25)
    elif region in ("vallée du rhône", "rhône"):
        window = (5, 18)
    elif region == "champagne":
        return "À boire maintenant"
    elif region == "alsace":
        if "grand cru" in appellation:
            window = (8, 25)
        else:
            window = (5, 15)
    elif region == "loire":
        if "demi-sec" in b.get("bouteille", "").lower() or "moelleux" in appellation:
            window = (10, 40)
        else:
            window = (5, 15)
    elif region in ("espagne", "italie"):
        if any(w in appellation for w in ["ribera", "priorat", "barolo", "brunello", "toscana"]):
            window = (10, 30)
        else:
            window = (5, 18)
    elif region == "liban":
        window = (8, 30)
    else:
        return "À évaluer"

    # Ajuster selon l'état
    if etat in ("Bon", "Moyen", "Abîmé", "À vérifier"):
        window = (window[0], max(window[0] + 3, window[1] - 5))

    start = mill + window[0]
    end = mill + window[1]
    current_year = 2026

    if end < current_year:
        return "À boire maintenant"

    return f"{start}–{end}"


def normaliser_format(fmt):
    """Convertit tout format en cl. Ex: '1,5L' → '150cl', '0,375' → '37,5cl', '62' → '62cl'."""
    fmt = fmt.strip()
    if fmt.lower().endswith("cl"):
        return fmt
    # Déterminer si le nombre est en litres ou déjà en cl
    has_l = "l" in fmt.lower()
    cleaned = fmt.lower().replace("l", "").replace(",", ".").strip()
    if not cleaned:
        return "75cl"
    try:
        val = float(cleaned)
        # Si suffixe "L" explicite → litres. Sinon : >= 10 = cl, < 10 = litres
        if has_l or val < 10:
            cl = val * 100
        else:
            cl = val
        if cl == int(cl):
            return f"{int(cl)}cl"
        else:
            return f"{cl:g}cl".replace(".", ",")
    except ValueError:
        return fmt


# Normaliser les formats en cl + auto-calcul de l'apogée
for b in INVENTAIRE:
    b["format"] = normaliser_format(b.get("format", "75cl"))
    if not b.get("apogee") or b["apogee"] == "":
        b["apogee"] = estimate_apogee(b)


# ──────────────────────────────────────────────
# GÉNÉRATION DES GRAPHIQUES (Matplotlib)
# ──────────────────────────────────────────────
def make_chart_image(fig, dpi=200):
    """Sauvegarde un figure matplotlib dans un buffer."""
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight',
                facecolor='#F2EEE2', edgecolor='none', transparent=False)
    plt.close(fig)
    buf.seek(0)
    return buf


def safe_chart_image(chart_buf, width, height):
    """Retourne un RLImage ou un Paragraph placeholder si le graphique est None."""
    if chart_buf is None:
        return Paragraph("<i>[Graphique non disponible]</i>",
                         ParagraphStyle('ChartNA', fontName='Helvetica-Oblique',
                                        fontSize=9, textColor=HexColor("#999999"),
                                        alignment=TA_CENTER, spaceBefore=10*mm, spaceAfter=10*mm))
    return RLImage(chart_buf, width=width, height=height)


def _group_small_slices(data_dict, threshold_pct=5):
    """Regroupe les entrées < threshold_pct% dans 'Autres'."""
    total = sum(data_dict.values())
    if total == 0:
        return data_dict
    grouped = {}
    autres = 0
    for k, v in sorted(data_dict.items(), key=lambda x: x[1], reverse=True):
        if (v / total * 100) >= threshold_pct:
            grouped[k] = v
        else:
            autres += v
    if autres > 0:
        grouped["Autres"] = autres
    return grouped


def create_region_chart():
    """Répartition par région (volume)."""
    try:
        regions = {}
        for b in INVENTAIRE:
            r = b["region"]
            regions[r] = regions.get(r, 0) + b["qte"]

        regions = _group_small_slices(regions)
        labels = list(regions.keys())
        sizes = list(regions.values())
        colors_list = ['#722F37', '#C5A258', '#A0616D', '#2D1B2E', '#D4841A',
                       '#2E7D4F', '#8B6F47', '#5C4033', '#B33A3A', '#4A6741']

        fig, ax = plt.subplots(figsize=(7, 5))
        wedges, texts, autotexts = ax.pie(
            sizes, labels=None, autopct='%1.0f%%',
            colors=colors_list[:len(labels)],
            startangle=90, pctdistance=0.75,
            wedgeprops=dict(width=0.45, edgecolor='white', linewidth=2)
        )
        for t in autotexts:
            t.set_fontsize(10)
            t.set_color('white')
            t.set_fontweight('bold')

        ax.legend(labels, loc='center left', bbox_to_anchor=(1.05, 0.5),
                  fontsize=9, frameon=False)
        ax.set_title("Répartition des bouteilles par région", fontsize=12, fontweight='bold',
                     color='#2D1B2E', pad=15)
        try:
            fig.tight_layout()
        except ValueError:
            pass
        return make_chart_image(fig)
    except Exception as e:
        print(f"Avertissement : graphique région (volume) ignoré — {e}")
        plt.close('all')
        return None


def create_color_chart():
    """Répartition par couleur/type de vin."""
    try:
        colors_dict = {}
        for b in INVENTAIRE:
            c = b["couleur"]
            colors_dict[c] = colors_dict.get(c, 0) + b["qte"]

        labels = [k for k, v in colors_dict.items() if v > 0]
        sizes = [colors_dict[k] for k in labels]
        cmap = {"Rouge": "#722F37", "Rosé": "#E8A0B0", "Blanc": "#E8D5A3",
                "Effervescent": "#C5A258", "Liquoreux": "#D4841A"}
        colors_list = [cmap.get(l, "#888") for l in labels]

        fig, ax = plt.subplots(figsize=(7, 5))
        wedges, texts, autotexts = ax.pie(
            sizes, labels=None, autopct='%1.0f%%',
            colors=colors_list,
            startangle=90, pctdistance=0.75,
            wedgeprops=dict(width=0.45, edgecolor='white', linewidth=2)
        )
        for t in autotexts:
            t.set_fontsize(10)
            t.set_fontweight('bold')
            t.set_color('#2D1B2E')

        ax.legend(labels, loc='center left', bbox_to_anchor=(1.05, 0.5),
                  fontsize=9, frameon=False)
        ax.set_title("Répartition par couleur", fontsize=12, fontweight='bold',
                     color='#2D1B2E', pad=15)
        try:
            fig.tight_layout()
        except ValueError:
            pass
        return make_chart_image(fig)
    except Exception as e:
        print(f"Avertissement : graphique couleur ignoré — {e}")
        plt.close('all')
        return None


def create_recommendation_chart():
    """Répartition des recommandations (garder/vendre/attendre)."""
    try:
        reco_count = {"À conserver": 0, "À vendre": 0, "À surveiller": 0}
        reco_value = {"À conserver": 0, "À vendre": 0, "À surveiller": 0}
        for b in INVENTAIRE:
            r = b["reco"]
            reco_count[r] = reco_count.get(r, 0) + b["qte"]
            reco_value[r] = reco_value.get(r, 0) + b["qte"] * b["val_unit"]

        categories = list(reco_count.keys())
        counts = [reco_count[c] for c in categories]
        values = [reco_value[c] for c in categories]
        colors_bars = ['#2E7D4F', '#B33A3A', '#D4841A']

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 6))

        bars1 = ax1.bar(categories, counts, color=colors_bars, edgecolor='white', linewidth=1, width=0.45)
        ax1.set_title("Par nombre de bouteilles", fontsize=11, fontweight='bold', color='#2D1B2E')
        ax1.spines['top'].set_visible(False)
        ax1.spines['right'].set_visible(False)
        ax1.spines['left'].set_color('#E0D8CC')
        ax1.spines['bottom'].set_color('#E0D8CC')
        ax1.tick_params(colors='#4A4A4A', labelsize=11)
        max_count = max(counts) if counts else 1
        max_count = max(max_count, 1)  # Éviter ylim identique si tout est à 0
        ax1.set_ylim(0, max_count * 1.15)
        for bar, val in zip(bars1, counts):
            ax1.text(bar.get_x() + bar.get_width()/2., bar.get_height() + max_count * 0.02,
                     str(val), ha='center', va='bottom', fontsize=11, fontweight='bold', color='#2D1B2E')

        bars2 = ax2.bar(categories, values, color=colors_bars, edgecolor='white', linewidth=1, width=0.45)
        ax2.set_title("Par valeur estimée (€)", fontsize=11, fontweight='bold', color='#2D1B2E')
        ax2.spines['top'].set_visible(False)
        ax2.spines['right'].set_visible(False)
        ax2.spines['left'].set_color('#E0D8CC')
        ax2.spines['bottom'].set_color('#E0D8CC')
        ax2.tick_params(colors='#4A4A4A', labelsize=11)
        ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{int(x):,}'.replace(',', ' ')))
        max_val = max(values) if values else 1
        max_val = max(max_val, 1)
        ax2.set_ylim(0, max_val * 1.15)
        for bar, val in zip(bars2, values):
            ax2.text(bar.get_x() + bar.get_width()/2., bar.get_height() + max_val * 0.02,
                     f'{val:,}€'.replace(',', ' '), ha='center', va='bottom', fontsize=10, fontweight='bold', color='#2D1B2E')

        fig.suptitle("Recommandations", fontsize=12, fontweight='bold', color='#2D1B2E', y=0.98)
        try:
            fig.tight_layout(rect=[0, 0, 1, 0.95])
        except ValueError:
            pass
        return make_chart_image(fig)
    except Exception as e:
        print(f"Avertissement : graphique recommandations ignoré — {e}")
        plt.close('all')
        return None


def create_apogee_chart():
    """Potentiel de garde."""
    try:
        current_year = 2026
        categories = {"À boire maintenant": 0, "Proche apogée (1–3 ans)": 0,
                      "En développement (3–10 ans)": 0, "Garde longue (10+ ans)": 0}
        for b in INVENTAIRE:
            apogee = b["apogee"]
            if apogee == "À boire":
                categories["À boire maintenant"] += b["qte"]
            else:
                try:
                    parts = apogee.replace("–", "-").split("-")
                    start = int(parts[0])
                    end = int(parts[1])
                    mid = (start + end) / 2
                    diff = mid - current_year
                    if diff <= 0:
                        categories["À boire maintenant"] += b["qte"]
                    elif diff <= 3:
                        categories["Proche apogée (1–3 ans)"] += b["qte"]
                    elif diff <= 10:
                        categories["En développement (3–10 ans)"] += b["qte"]
                    else:
                        categories["Garde longue (10+ ans)"] += b["qte"]
                except:
                    categories["À boire maintenant"] += b["qte"]

        labels = [k for k, v in categories.items() if v > 0]
        sizes = [categories[k] for k in labels]
        colors_list = ['#B33A3A', '#D4841A', '#C5A258', '#2E7D4F'][:len(labels)]

        fig, ax = plt.subplots(figsize=(10, 5))
        bars = ax.barh(labels, sizes, color=colors_list, edgecolor='white', linewidth=1, height=0.5)
        ax.set_title("Potentiel de garde", fontsize=12, fontweight='bold', color='#2D1B2E', pad=15)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#E0D8CC')
        ax.spines['bottom'].set_color('#E0D8CC')
        ax.tick_params(colors='#4A4A4A', labelsize=11)
        ax.invert_yaxis()
        max_size = max(sizes) if sizes else 1
        max_size = max(max_size, 1)
        ax.set_xlim(0, max_size * 1.15)
        for bar, val in zip(bars, sizes):
            ax.text(bar.get_width() + max_size * 0.02, bar.get_y() + bar.get_height()/2.,
                    f'{val} btls', ha='left', va='center', fontsize=11, fontweight='bold', color='#2D1B2E')
        try:
            fig.tight_layout()
        except ValueError:
            pass
        return make_chart_image(fig)
    except Exception as e:
        print(f"Avertissement : graphique potentiel de garde ignoré — {e}")
        plt.close('all')
        return None


def create_region_value_chart():
    """Camembert poids financier par région."""
    try:
        regions = {}
        for b in INVENTAIRE:
            r = b["region"]
            regions[r] = regions.get(r, 0) + b["qte"] * b["val_unit"]

        regions = _group_small_slices(regions)
        labels = list(regions.keys())
        sizes = list(regions.values())
        colors_list = ['#722F37', '#C5A258', '#A0616D', '#2D1B2E', '#D4841A',
                       '#2E7D4F', '#8B6F47', '#5C4033', '#B33A3A', '#4A6741']

        fig, ax = plt.subplots(figsize=(7, 5))
        wedges, texts, autotexts = ax.pie(
            sizes, labels=None, autopct='%1.0f%%',
            colors=colors_list[:len(labels)],
            startangle=90, pctdistance=0.75,
            wedgeprops=dict(width=0.45, edgecolor='white', linewidth=2)
        )
        for t in autotexts:
            t.set_fontsize(10)
            t.set_color('white')
            t.set_fontweight('bold')

        ax.legend(labels, loc='center left', bbox_to_anchor=(1.05, 0.5),
                  fontsize=9, frameon=False)
        ax.set_title("Valeur estimée par région (€)", fontsize=12, fontweight='bold',
                     color='#2D1B2E', pad=15)
        try:
            fig.tight_layout()
        except ValueError:
            pass
        return make_chart_image(fig)
    except Exception as e:
        print(f"Avertissement : graphique valeur par région ignoré — {e}")
        plt.close('all')
        return None


# ──────────────────────────────────────────────
# CONSTRUCTION DU PDF
# ──────────────────────────────────────────────
MOIS_FR = {1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril", 5: "Mai", 6: "Juin",
           7: "Juillet", 8: "Août", 9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"}

def generate_output_path():
    """Génère le nom du fichier PDF selon le format :
    Audit et estimation de cave - NOM - REF - Mois Année.pdf"""
    nom = CLIENT_DATA.get("nom", "Client")
    # Supprimer titres de civilité et extraire le nom de famille
    cleaned = nom.replace("M.", "").replace("Mme", "").replace("et", "").strip()
    words = [w for w in cleaned.split() if len(w) > 1]
    nom_famille = words[-1].upper() if words else "CLIENT"
    # Supprimer accents pour le nom de fichier
    normalized = unicodedata.normalize('NFD', nom_famille)
    nom_famille_ascii = normalized.encode('ascii', 'ignore').decode('ascii')

    ref = CLIENT_DATA.get("ref_mission", "EST-0000-00-XXX")
    now = datetime.now()
    mois_annee = f"{MOIS_FR.get(now.month, 'Mois')} {now.year}"

    filename = f"Audit et estimation de cave - {nom_famille_ascii} - {ref} - {mois_annee}.pdf"
    return f"/Users/fannylonqueubrochard/Desktop/{filename}"

OUTPUT_PATH = generate_output_path()


LAND_W, LAND_H = A4[1], A4[0]  # Landscape dimensions


class NumberedCanvas(canvas.Canvas):
    """Canvas personnalisé qui gère le numéro total de pages (Page X/Y)."""
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_footer(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def _draw_footer(self, total_pages):
        """Pied de page uniforme sur toutes les pages (sauf couverture)."""
        page_num = self._pageNumber
        if page_num == 1:
            return  # Pas de footer sur la couverture

        self.saveState()

        # Déterminer la largeur de la page courante
        pw = float(self._pagesize[0])
        ph = float(self._pagesize[1])

        # Ligne dorée en haut
        self.setStrokeColor(GOLD)
        self.setLineWidth(1.5)
        self.line(MARGIN_L, ph - 14*mm, pw - MARGIN_R, ph - 14*mm)

        # Ligne dorée en bas
        self.setStrokeColor(GOLD)
        self.setLineWidth(0.5)
        self.line(MARGIN_L, 17*mm, pw - MARGIN_R, 17*mm)

        # Footer centré : CLIENT — REF — Rapport confidentiel — Page X/Y
        footer_text = (
            f"{CLIENT_DATA['nom']} — {CLIENT_DATA['ref_mission']} — "
            f"Rapport confidentiel — Page {page_num}/{total_pages}"
        )
        self.setFont("Helvetica", 7)
        self.setFillColor(TEXT_LIGHT)
        self.drawCentredString(pw / 2, 12*mm, footer_text)

        self.restoreState()


def draw_page_background(canvas_obj, doc):
    """Header uniquement — le footer est géré par NumberedCanvas."""
    pass


def draw_landscape_background(canvas_obj, doc):
    """Header uniquement — le footer est géré par NumberedCanvas."""
    pass


def draw_cover_page(canvas_obj, doc):
    """Page de garde — fond crème, logo cave centré, style premium épuré."""
    canvas_obj.saveState()

    # Couleurs couverture
    cover_bg = HexColor("#F2EEE2")
    cover_gold = HexColor("#C9A84C")
    cover_text = HexColor("#2D1B2E")
    cover_text_light = HexColor("#6B5F65")

    # Fond blanc (extérieur du liseré)
    canvas_obj.setFillColor(white)
    canvas_obj.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)

    # Fond crème (intérieur du liseré)
    canvas_obj.setFillColor(cover_bg)
    canvas_obj.rect(MARGIN_L, MARGIN_B, PAGE_W - MARGIN_L - MARGIN_R,
                    PAGE_H - MARGIN_T - MARGIN_B, fill=1, stroke=0)

    # ── Liseré doré — encadrement au niveau des marges ──
    canvas_obj.setStrokeColor(cover_gold)
    canvas_obj.setLineWidth(0.6)
    canvas_obj.rect(MARGIN_L, MARGIN_B, PAGE_W - MARGIN_L - MARGIN_R,
                    PAGE_H - MARGIN_T - MARGIN_B, fill=0, stroke=1)

    cx = PAGE_W / 2

    # ── Logo cave centré en haut ──
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cover_logo.png")
    if os.path.exists(logo_path):
        logo_w = 55 * mm
        logo_h = 55 * mm
        canvas_obj.drawImage(logo_path, cx - logo_w/2, PAGE_H - 85*mm,
                             width=logo_w, height=logo_h, mask='auto', preserveAspectRatio=True)

    # ── Filet doré sous le logo ──
    canvas_obj.setStrokeColor(cover_gold)
    canvas_obj.setLineWidth(0.4)
    rule_inset = 55 * mm
    canvas_obj.line(rule_inset, PAGE_H - 95*mm, PAGE_W - rule_inset, PAGE_H - 95*mm)

    # ── 3 filets : haut (sous logo), milieu (sous titre), bas ──
    top_rule_y = PAGE_H - 95*mm
    bottom_rule_y = 45*mm
    mid_rule_y = (top_rule_y + bottom_rule_y) / 2

    # ── Titre — centré entre filet haut et filet milieu ──
    # Bloc titre : 3 lignes (12mm entre chaque) + sous-titre = ~42mm de haut
    title_center_y = (top_rule_y + mid_rule_y) / 2
    canvas_obj.setFillColor(cover_text)
    canvas_obj.setFont("Helvetica-Bold", 20)
    canvas_obj.drawCentredString(cx, title_center_y + 12*mm, "RAPPORT D'AUDIT ET D'ESTIMATION")
    canvas_obj.drawCentredString(cx, title_center_y - 2*mm, "DE CAVE À VINS")

    # Sous-titre
    canvas_obj.setFillColor(cover_text_light)
    canvas_obj.setFont("Helvetica-Oblique", 9)
    canvas_obj.drawCentredString(cx, title_center_y - 18*mm,
        "Inventaire, valorisation de marché et recommandations d'arbitrage patrimonial")

    # ── Filet milieu (sous titre) ──
    canvas_obj.setStrokeColor(cover_gold)
    canvas_obj.setLineWidth(0.3)
    line_w = 30 * mm
    canvas_obj.line(cx - line_w, mid_rule_y, cx + line_w, mid_rule_y)

    # ── Infos client — centrées entre filet milieu et filet bas ──
    # Bloc client : "Établi pour" + nom + 3 lignes détails = ~45mm de haut
    client_center_y = (mid_rule_y + bottom_rule_y) / 2
    canvas_obj.setFillColor(cover_text_light)
    canvas_obj.setFont("Helvetica", 9)
    canvas_obj.drawCentredString(cx, client_center_y + 20*mm, "Établi pour")

    canvas_obj.setFillColor(cover_text)
    canvas_obj.setFont("Helvetica-Bold", 16)
    canvas_obj.drawCentredString(cx, client_center_y + 8*mm, CLIENT_DATA["nom"])

    canvas_obj.setFont("Helvetica", 8.5)
    canvas_obj.setFillColor(cover_text_light)
    canvas_obj.drawCentredString(cx, client_center_y - 6*mm, f"Référence : {CLIENT_DATA['ref_mission']}")
    canvas_obj.drawCentredString(cx, client_center_y - 14*mm, f"Date d'émission : {CLIENT_DATA['date_rapport']}")
    canvas_obj.drawCentredString(cx, client_center_y - 22*mm, f"Objectif : {CLIENT_DATA['objectif']}")

    # ── Filet bas ──
    canvas_obj.setStrokeColor(cover_gold)
    canvas_obj.setLineWidth(0.4)
    canvas_obj.line(rule_inset, 45*mm, PAGE_W - rule_inset, 45*mm)

    # ── Mention confidentiel ──
    canvas_obj.setFillColor(Color(0.42, 0.37, 0.40, alpha=0.6))
    canvas_obj.setFont("Helvetica", 6.5)
    canvas_obj.drawCentredString(cx, 36*mm,
        "Ce rapport est strictement confidentiel et destiné à l'usage exclusif du client désigné.")
    canvas_obj.drawCentredString(cx, 30*mm,
        CLIENT_DATA.get("contact_email", ""))

    canvas_obj.restoreState()


def build_styles():
    """Styles de paragraphes personnalisés."""
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=15,
        textColor=WINE_DARK,
        spaceAfter=4*mm,
        spaceBefore=6*mm,
        leading=18,
    ))
    styles.add(ParagraphStyle(
        'SectionSubtitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=WINE_PRIMARY,
        spaceAfter=3*mm,
        spaceBefore=4*mm,
        leading=14,
    ))
    styles.add(ParagraphStyle(
        'BodyText2',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=TEXT_DARK,
        leading=14,
        spaceAfter=2*mm,
        alignment=TA_JUSTIFY,
    ))
    styles.add(ParagraphStyle(
        'Intro',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        textColor=TEXT_MID,
        leading=14,
        spaceAfter=4*mm,
        spaceBefore=1*mm,
        alignment=TA_JUSTIFY,
    ))
    styles.add(ParagraphStyle(
        'SmallNote',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7.5,
        textColor=TEXT_LIGHT,
        leading=10,
        spaceAfter=2*mm,
    ))
    styles.add(ParagraphStyle(
        'KPI_Number',
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=WINE_PRIMARY,
        alignment=TA_CENTER,
        leading=26,
    ))
    styles.add(ParagraphStyle(
        'KPI_Label',
        fontName='Helvetica',
        fontSize=8,
        textColor=TEXT_MID,
        alignment=TA_CENTER,
        leading=10,
    ))
    return styles


def gold_divider():
    """Ligne de séparation dorée."""
    t = Table([['']],
              colWidths=[PAGE_W - MARGIN_L - MARGIN_R],
              rowHeights=[0.5*mm])
    t.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 0.5, GOLD),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    return t


def section_number_title(num, title):
    """Numéro de section + titre avec ligne dorée."""
    return Paragraph(
        f'<font color="{GOLD.hexval()}" size="18">{num}.</font>&nbsp;&nbsp;'
        f'<font color="{WINE_DARK.hexval()}" size="15"><b>{title}</b></font>',
        ParagraphStyle('SecNum', fontName='Helvetica-Bold', fontSize=15,
                       leading=20, spaceAfter=2*mm, spaceBefore=5*mm)
    )


def build_report():
    """Construction du rapport complet."""
    styles = build_styles()

    doc = BaseDocTemplate(
        OUTPUT_PATH,
        pagesize=A4,
        leftMargin=MARGIN_L,
        rightMargin=MARGIN_R,
        topMargin=MARGIN_T,
        bottomMargin=MARGIN_B,
        title="Rapport d'estimation de cave",
        author="Service d'estimation de cave",
    )

    # Templates de page
    cover_template = PageTemplate(
        id='cover',
        frames=[Frame(0, 0, PAGE_W, PAGE_H, id='cover_frame',
                       leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)],
        onPage=draw_cover_page,
    )
    content_template = PageTemplate(
        id='content',
        frames=[Frame(MARGIN_L, MARGIN_B, PAGE_W - MARGIN_L - MARGIN_R,
                       PAGE_H - MARGIN_T - MARGIN_B, id='content_frame')],
        onPage=draw_page_background,
    )
    landscape_template = PageTemplate(
        id='landscape',
        frames=[Frame(MARGIN_L, MARGIN_B, LAND_W - MARGIN_L - MARGIN_R,
                       LAND_H - MARGIN_T - MARGIN_B, id='landscape_frame')],
        onPage=draw_landscape_background,
        pagesize=(LAND_W, LAND_H),
    )
    doc.addPageTemplates([cover_template, content_template, landscape_template])

    story = []

    # ═══ PAGE DE GARDE ═══
    from reportlab.platypus import NextPageTemplate
    story.append(NextPageTemplate('content'))
    story.append(PageBreak())

    # ═══ SYNTHÈSE EXÉCUTIVE ═══
    content_w = PAGE_W - MARGIN_L - MARGIN_R
    pad = 4*mm  # ~15px inner padding

    # Titre
    story.append(Paragraph(
        f'<font color="{GOLD.hexval()}" size="18">—</font>&nbsp;&nbsp;'
        f'<font color="{WINE_DARK.hexval()}" size="16"><b>SYNTHÈSE EXÉCUTIVE</b></font>'
        f'&nbsp;&nbsp;<font color="{GOLD.hexval()}" size="18">—</font>',
        ParagraphStyle('ExecTitle', fontName='Helvetica-Bold', fontSize=16,
                       alignment=TA_CENTER, spaceAfter=5*mm, spaceBefore=2*mm, leading=20)))
    story.append(gold_divider())
    story.append(Spacer(1, 5*mm))

    # ── Calculs automatiques ──
    reco_stats = {"À conserver": {"lots": 0, "btls": 0, "val": 0},
                  "À vendre": {"lots": 0, "btls": 0, "val": 0},
                  "À surveiller": {"lots": 0, "btls": 0, "val": 0}}
    for b in INVENTAIRE:
        r = b["reco"]
        reco_stats[r]["lots"] += 1
        reco_stats[r]["btls"] += b["qte"]
        reco_stats[r]["val"] += b["qte"] * b["val_unit"]

    low_est = math.ceil(total_val() * 0.85 / 5) * 5
    high_est = math.ceil(total_val() * 1.10 / 5) * 5

    # Concentration de valeur : top 3
    sorted_by_val = sorted(INVENTAIRE, key=lambda x: x["qte"] * x["val_unit"], reverse=True)
    top3_val = sum(b["qte"] * b["val_unit"] for b in sorted_by_val[:3])
    top3_pct = int(top3_val / total_val() * 100) if total_val() > 0 else 0

    # Dominante régionale — volume
    reg_count = {}
    for b in INVENTAIRE:
        reg_count[b["region"]] = reg_count.get(b["region"], 0) + b["qte"]
    dominant_vol_region = max(reg_count, key=reg_count.get) if reg_count else "N/A"
    dominant_vol_pct = int(reg_count.get(dominant_vol_region, 0) / total_btls() * 100) if total_btls() > 0 else 0

    # Dominante régionale — valeur
    reg_val = {}
    for b in INVENTAIRE:
        reg_val[b["region"]] = reg_val.get(b["region"], 0) + b["qte"] * b["val_unit"]
    dominant_val_region = max(reg_val, key=reg_val.get) if reg_val else "N/A"
    dominant_val_pct = int(reg_val.get(dominant_val_region, 0) / total_val() * 100) if total_val() > 0 else 0

    # Liquidité — depuis CLIENT_DATA
    liquidite = CLIENT_DATA.get("liquidite_cave", "Moyenne")

    # Conservation
    etat_scores = {"Excellent": 4, "Très bon": 3, "Bon": 2, "Moyen": 1, "Abîmé": 0, "À vérifier": 1}
    avg_etat = sum(etat_scores.get(b["etat"], 2) * b["qte"] for b in INVENTAIRE) / total_btls() if total_btls() > 0 else 2
    if avg_etat >= 3.5:
        conservation = "Excellent état apparent"
    elif avg_etat >= 2.5:
        conservation = "Bon état apparent"
    else:
        conservation = "État moyen — vigilance"

    # ── Anneau graphique (légende à côté, pas dessus) ──
    reco_fig = plt.figure(figsize=(3.2, 1.8))
    gs = reco_fig.add_gridspec(1, 2, width_ratios=[1, 0.6], wspace=0.05)
    reco_ax = reco_fig.add_subplot(gs[0])
    leg_ax = reco_fig.add_subplot(gs[1])
    leg_ax.axis('off')

    reco_labels = []
    reco_sizes = []
    reco_colors_chart = []
    color_map_chart = {"À conserver": "#2E7D4F", "À vendre": "#B33A3A", "À surveiller": "#D4841A"}
    for rname in ["À conserver", "À vendre", "À surveiller"]:
        if reco_stats[rname]["val"] > 0:
            reco_labels.append(rname)
            reco_sizes.append(reco_stats[rname]["val"])
            reco_colors_chart.append(color_map_chart[rname])

    total_reco_val = sum(reco_sizes)
    reco_pcts = [v / total_reco_val * 100 for v in reco_sizes] if total_reco_val > 0 else [0]*len(reco_sizes)

    wedges, texts = reco_ax.pie(
        reco_sizes, labels=None,
        colors=reco_colors_chart, startangle=90,
        wedgeprops=dict(width=0.38, edgecolor='white', linewidth=2))

    # Légende à droite avec pourcentages
    for i, (label, color, pct) in enumerate(zip(reco_labels, reco_colors_chart, reco_pcts)):
        y_pos = 0.65 - i * 0.3
        leg_ax.add_patch(plt.Rectangle((0.05, y_pos - 0.06), 0.15, 0.12,
                                        facecolor=color, edgecolor='none'))
        leg_ax.text(0.28, y_pos, f'{label} {pct:.0f}%', fontsize=8, va='center',
                    fontweight='bold', color='#1A1A1A')
    leg_ax.set_xlim(0, 1)
    leg_ax.set_ylim(-0.1, 1)

    reco_fig.subplots_adjust(left=0.05, right=0.95, top=0.95, bottom=0.05)
    reco_chart_buf = make_chart_image(reco_fig, dpi=200)

    # ══════════════════════════════════════════════
    # BLOC 1 — VUE D'ENSEMBLE (pleine largeur)
    # ══════════════════════════════════════════════
    val_str = f'{total_val():,}€'.replace(",", " ")
    fourch_str = f'{low_est:,}€ — {high_est:,}€'.replace(",", " ")

    kpi_main = Paragraph(
        f'<font size="30" color="{WINE_PRIMARY.hexval()}"><b>{val_str}</b></font>',
        ParagraphStyle('KPIMain', alignment=TA_LEFT, leading=36))
    kpi_fourch = Paragraph(
        f'<font size="8" color="{TEXT_MID.hexval()}">Fourchette indicative : </font>'
        f'<font size="8" color="{WINE_PRIMARY.hexval()}"><b>{fourch_str}</b></font>',
        ParagraphStyle('KPIFourch', alignment=TA_LEFT, leading=12, spaceBefore=1*mm))
    kpi_details = Paragraph(
        f'<font size="18" color="{WINE_PRIMARY.hexval()}"><b>{len(INVENTAIRE)}</b></font>'
        f'<font size="9" color="{TEXT_MID.hexval()}"> références</font>'
        f'&nbsp;&nbsp;&nbsp;&nbsp;'
        f'<font size="18" color="{WINE_PRIMARY.hexval()}"><b>{total_btls()}</b></font>'
        f'<font size="9" color="{TEXT_MID.hexval()}"> bouteilles</font>',
        ParagraphStyle('KPIDetails', alignment=TA_LEFT, leading=22, spaceBefore=3*mm))

    left_kpi = Table([[kpi_main], [kpi_fourch], [kpi_details]],
                     colWidths=[content_w * 0.58])
    left_kpi.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), pad),
        ('TOPPADDING', (0, 0), (0, 0), pad),
        ('BOTTOMPADDING', (0, -1), (0, -1), pad),
    ]))

    ring_img = RLImage(reco_chart_buf, width=52*mm, height=30*mm)
    right_ring = Table([[ring_img]], colWidths=[content_w * 0.42])
    right_ring.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    bloc1_title = Paragraph(
        f'<font color="{WINE_PRIMARY.hexval()}"><b>VUE D\'ENSEMBLE</b></font>',
        ParagraphStyle('B1Title', fontName='Helvetica-Bold', fontSize=9,
                       spaceBefore=1*mm, spaceAfter=1*mm))

    bloc1_content = Table([[left_kpi, right_ring]],
                          colWidths=[content_w * 0.58, content_w * 0.42])
    bloc1_content.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    bloc1 = Table([[bloc1_title], [bloc1_content]],
                  colWidths=[content_w])
    bloc1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_LIGHT),
        ('LINEBELOW', (0, 0), (0, 0), 0.8, GOLD),
        ('LEFTPADDING', (0, 0), (-1, -1), pad),
        ('RIGHTPADDING', (0, 0), (-1, -1), pad),
        ('TOPPADDING', (0, 0), (-1, -1), 2*mm),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 2*mm),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(bloc1)
    story.append(Spacer(1, 4*mm))

    # ══════════════════════════════════════════════
    # BLOC 2 — DIAGNOSTIC CAVE (pleine largeur)
    # ══════════════════════════════════════════════
    diag_items = [
        ["Dominante régionale (volume)", f"{dominant_vol_region} — {dominant_vol_pct}% des bouteilles"],
        ["Dominante régionale (valeur)", f"{dominant_val_region} — {dominant_val_pct}% de la valeur"],
        ["Concentration de valeur", f"{top3_pct}% de la valeur concentrée dans 3 références"],
        ["Liquidité globale", liquidite],
        ["Conservation apparente", conservation],
    ]

    diag_rows = []
    for label, value in diag_items:
        diag_rows.append([
            Paragraph(f'<font size="8" color="{TEXT_MID.hexval()}"><b>{label}</b></font>',
                      ParagraphStyle('DL', leading=11)),
            Paragraph(f'<font size="9" color="{TEXT_DARK.hexval()}">{value}</font>',
                      ParagraphStyle('DV', leading=12)),
        ])

    diag_table = Table(diag_rows,
                       colWidths=[content_w * 0.35, content_w * 0.58])
    diag_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, -2), 0.2, BORDER_LIGHT),
        ('LEFTPADDING', (0, 0), (-1, -1), 2*mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2*mm),
    ]))

    bloc2_title = Paragraph(
        f'<font color="{WINE_PRIMARY.hexval()}"><b>DIAGNOSTIC CAVE</b></font>',
        ParagraphStyle('B2Title', fontName='Helvetica-Bold', fontSize=9,
                       spaceBefore=1*mm, spaceAfter=1*mm))

    bloc2 = Table([[bloc2_title], [diag_table]],
                  colWidths=[content_w])
    bloc2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_LIGHT),
        ('LINEBELOW', (0, 0), (0, 0), 0.8, GOLD),
        ('LEFTPADDING', (0, 0), (-1, -1), pad),
        ('RIGHTPADDING', (0, 0), (-1, -1), pad),
        ('TOPPADDING', (0, 0), (-1, -1), 2*mm),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 2*mm),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(bloc2)
    story.append(Spacer(1, 4*mm))

    # ══════════════════════════════════════════════
    # BLOC 3 — RECOMMANDATION SYNTHÉTIQUE (pleine largeur)
    # ══════════════════════════════════════════════
    reco_header = ["", "Lots", "Bouteilles", "Valeur estimée"]
    reco_rows_data = [reco_header]
    reco_display = [
        ("À conserver", GREEN_REC), ("À vendre", RED_REC), ("À surveiller", ORANGE_REC)]
    for rname, rcolor in reco_display:
        s = reco_stats[rname]
        reco_rows_data.append([
            rname, str(s["lots"]), str(s["btls"]),
            f'{s["val"]:,}€'.replace(",", " ")])

    inner_w = content_w - 2 * pad  # Width available inside bloc padding
    reco_synth = Table(reco_rows_data,
                       colWidths=[inner_w * 0.30, inner_w * 0.20,
                                  inner_w * 0.20, inner_w * 0.30],
                       rowHeights=[8*mm] * 4)
    reco_ts3 = [
        ('BACKGROUND', (0, 0), (-1, 0), BG_LIGHT),
        ('TEXTCOLOR', (0, 0), (-1, 0), WINE_DARK),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TEXTCOLOR', (0, 1), (-1, -1), TEXT_DARK),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, BG_LIGHT]),
        ('LINEBELOW', (0, 0), (-1, 0), 0.6, GOLD),
        ('LINEBELOW', (0, 1), (-1, -1), 0.2, BORDER_LIGHT),
        ('LEFTPADDING', (0, 0), (-1, -1), pad),
        ('RIGHTPADDING', (0, 0), (-1, -1), pad),
    ]
    for i, (rname, rcolor) in enumerate(reco_display, 1):
        reco_ts3.append(('TEXTCOLOR', (0, i), (0, i), rcolor))
        reco_ts3.append(('FONTNAME', (0, i), (0, i), 'Helvetica-Bold'))
    reco_synth.setStyle(TableStyle(reco_ts3))

    bloc3_title = Paragraph(
        f'<font color="{WINE_PRIMARY.hexval()}"><b>RECOMMANDATION SYNTHÉTIQUE</b></font>',
        ParagraphStyle('B3Title', fontName='Helvetica-Bold', fontSize=9,
                       spaceBefore=1*mm, spaceAfter=1*mm))

    bloc3 = Table([[bloc3_title], [reco_synth]],
                  colWidths=[content_w])
    bloc3.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_LIGHT),
        ('LINEBELOW', (0, 0), (0, 0), 0.8, GOLD),
        ('LEFTPADDING', (0, 0), (-1, -1), pad),
        ('RIGHTPADDING', (0, 0), (-1, -1), pad),
        ('TOPPADDING', (0, 0), (-1, -1), 2*mm),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 2*mm),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(bloc3)

    story.append(PageBreak())

    # ═══ SECTION 1 — INFORMATIONS MISSION ═══
    story.append(section_number_title("1", "INFORMATIONS DE LA MISSION"))
    story.append(gold_divider())
    story.append(Spacer(1, 2*mm))

    info_data = [
        ["Client", CLIENT_DATA["nom"]],
        ["Contact", CLIENT_DATA["email"]],
        ["Référence", CLIENT_DATA["ref_mission"]],
        ["Date du rapport", CLIENT_DATA["date_rapport"]],
        ["Objectif", CLIENT_DATA["objectif"]],
        ["Conditions de l'estimation", CLIENT_DATA.get("conditions_estimation", "[À compléter]")],
        ["Conditions de conservation", CLIENT_DATA.get("conditions_conservation", "[À compléter]")],
        ["Provenance déclarée", CLIENT_DATA.get("provenance", "[À compléter]")],
    ]
    loc = CLIENT_DATA.get("localisation", "").strip()
    if loc:
        info_data.append(["Localisation", loc])
    infos_comp = CLIENT_DATA.get("infos_complementaires", "").strip()
    infos_comp_idx = None
    if infos_comp:
        infos_comp_idx = len(info_data)
        info_data.append(["Informations complémentaires",
                          Paragraph(infos_comp, ParagraphStyle('InfoComp', fontName='Helvetica',
                                    fontSize=8.5, textColor=TEXT_DARK, leading=11))])
    info_data.append(["Nombre de références", str(len(INVENTAIRE))])
    info_data.append(["Nombre total de bouteilles", str(total_btls())])

    # Hauteurs : 7mm pour toutes les lignes, None pour infos_comp (auto)
    row_heights = [7*mm] * len(info_data)
    if infos_comp_idx is not None:
        row_heights[infos_comp_idx] = None

    info_table = Table(info_data,
                       colWidths=[55*mm, PAGE_W - MARGIN_L - MARGIN_R - 55*mm],
                       rowHeights=row_heights)
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('TEXTCOLOR', (0, 0), (0, -1), WINE_PRIMARY),
        ('TEXTCOLOR', (1, 0), (1, -1), TEXT_DARK),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [white, BG_LIGHT]),
        ('LINEBELOW', (0, 0), (-1, -1), 0.3, BORDER_LIGHT),
        ('LEFTPADDING', (0, 0), (-1, -1), 4*mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4*mm),
        ('TOPPADDING', (0, 0), (-1, -1), 1.5*mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5*mm),
    ]))
    story.append(info_table)

    # Méthodologie et limites
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph("<b>Méthodologie et limites de l'estimation</b>", styles['SectionSubtitle']))
    story.append(Paragraph(
        "Les estimations présentées dans ce rapport ont été établies selon une méthodologie croisée, "
        "fondée sur l'identification des références, l'observation de l'état apparent des flacons, "
        "les prix observés sur le marché secondaire et les résultats récents de ventes publiques comparables.",
        styles['BodyText2']))
    story.append(Spacer(1, 1*mm))
    story.append(Paragraph(
        "Sauf mention contraire, le présent rapport :<br/>"
        "— ne constitue pas une expertise judiciaire ;<br/>"
        "— ne vaut pas certification d'authenticité ;<br/>"
        "— ne garantit pas l'état organoleptique du vin ;<br/>"
        "— ne peut se substituer à une inspection physique approfondie ;<br/>"
        "— n'emporte aucune garantie de prix de vente effectif ;<br/>"
        "— doit être lu comme une estimation de marché à date, sous réserve de conformité "
        "entre les éléments transmis et les flacons réels.",
        styles['BodyText2']))
    story.append(Spacer(1, 1*mm))
    story.append(Paragraph(
        "Ce document constitue une estimation indicative de valeur marchande au jour de son émission. "
        "Il vise à fournir au client une lecture structurée de la cave auditée, ainsi qu'une aide "
        "à la décision dans une logique de conservation, d'arbitrage ou de cession partielle.",
        styles['BodyText2']))

    # Échelle de lecture
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph("<b>Échelle de lecture — état apparent des flacons</b>", styles['SectionSubtitle']))
    story.append(Paragraph(
        "— <b>Excellent</b> : présentation optimale, niveau satisfaisant, défauts absents ou négligeables<br/>"
        "— <b>Très bon</b> : légers signes du temps sans incidence majeure sur la désirabilité<br/>"
        "— <b>Bon</b> : présentation correcte, quelques défauts compatibles avec une valorisation normale<br/>"
        "— <b>Moyen</b> : défauts visibles susceptibles d'affecter la valeur ou la liquidité<br/>"
        "— <b>Abîmé</b> : dégradation significative du flacon, valeur et liquidité pouvant être fortement impactées<br/>"
        "— <b>À vérifier</b> : informations incomplètes ou état nécessitant un contrôle complémentaire",
        styles['BodyText2']))

    # ═══ SECTION 2 — INVENTAIRE DÉTAILLÉ (PAYSAGE) ═══
    story.append(NextPageTemplate('landscape'))
    story.append(PageBreak())

    story.append(section_number_title("2", "INVENTAIRE DÉTAILLÉ"))
    story.append(gold_divider())
    story.append(Spacer(1, 1.5*mm))
    story.append(Paragraph(
        "Chaque référence a été identifiée à partir de la liste et/ou des photos fournies. "
        "Les valeurs unitaires s'appuient sur les résultats d'enchères récents et les prix marché observés. "
        "L'inventaire détaillé constitue la base technique du présent rapport. "
        "Les pages suivantes en proposent une lecture synthétique, patrimoniale et opérationnelle.",
        styles['Intro']))

    # En-tête du tableau — 13 colonnes, avec Couleur
    inv_header = ["#", "Domaine / Cuvée", "Appellation", "Région", "Couleur",
                  "Format", "Mill.", "État", "CBO", "Qté", "Val. unit.", "Val. totale", "Orientation"]
    inv_rows = [inv_header]

    # Tri pour affichage : millésime croissant, NM à la fin, alpha en cas d'égalité
    inv_sorted = sorted(INVENTAIRE, key=lambda b: (
        0 if b["millesime"] != 0 else 1,   # NM (0) en dernier
        b["millesime"] if b["millesime"] != 0 else 9999,
        b["bouteille"].lower()
    ))
    inv_display = inv_sorted
    inv_text_style = ParagraphStyle('InvCell', fontName='Helvetica', fontSize=7,
                                     textColor=TEXT_DARK, leading=8.5)
    for i, b in enumerate(inv_display, 1):
        mill_str = str(b["millesime"]) if b["millesime"] else "NM"
        if b["millesime"] == 0:
            mill_str = "NM"
        val_t = b["qte"] * b["val_unit"]
        inv_rows.append([
            str(i),
            Paragraph(b["bouteille"], inv_text_style),
            Paragraph(b["appellation"], inv_text_style),
            b["region"],
            b["couleur"],
            b.get("format", "75cl"),
            mill_str,
            b["etat"],
            b.get("cbo", "Non"),
            str(b["qte"]),
            f'{b["val_unit"]}€',
            f'{val_t:,}€'.replace(",", " "),
            b["reco"],
        ])

    # Ligne totaux (13 colonnes) — toujours sur le total réel
    inv_rows.append(["", "", "", "", "", "", "", "", "", str(total_btls()), "",
                     f'{total_val():,}€'.replace(",", " "), ""])

    # Ligne légende (intégrée au tableau pour éviter qu'elle se retrouve seule sur une page)
    legend_text = Paragraph(
        "NM = Non millésimé · CBO = Caisse Bois d'Origine · "
        "Couleur : Rouge / Rosé / Blanc / Effervescent / Liquoreux · "
        "Les valeurs sont exprimées en euros (€) · "
        "État évalué sur photos : Excellent / Très bon / Bon / Moyen / Abîmé / À vérifier",
        ParagraphStyle('LegendInTable', fontName='Helvetica', fontSize=6,
                       textColor=TEXT_LIGHT, leading=8))
    inv_rows.append([legend_text] + [""] * 12)

    # Largeurs paysage — 13 colonnes
    col_widths = [7*mm, 50*mm, 46*mm, 22*mm, 19*mm, 13*mm, 12*mm, 15*mm, 11*mm, 9*mm, 16*mm, 18*mm, 14*mm]
    # Row heights auto pour permettre le retour à la ligne sur Domaine/Appellation
    inv_table = Table(inv_rows, colWidths=col_widths, repeatRows=1)

    # Style du tableau
    ts = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), WINE_DARK),
        ('TEXTCOLOR', (0, 0), (-1, 0), GOLD_LIGHT),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TEXTCOLOR', (0, 1), (-1, -2), TEXT_DARK),
        # Alignements
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),    # #
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),    # Couleur
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),    # Format
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),    # Mill.
        ('ALIGN', (7, 1), (7, -1), 'CENTER'),    # État
        ('ALIGN', (8, 1), (8, -1), 'CENTER'),    # CBO
        ('ALIGN', (9, 1), (9, -1), 'CENTER'),    # Qté
        ('ALIGN', (10, 1), (11, -1), 'RIGHT'),   # Val. unit. + Val. totale
        ('ALIGN', (12, 1), (12, -1), 'CENTER'),  # Reco.
        # Alternance
        ('ROWBACKGROUNDS', (0, 1), (-1, -3), [white, BG_LIGHT]),
        ('LINEBELOW', (0, 0), (-1, -3), 0.2, BORDER_LIGHT),
        # Ligne total (avant-dernière ligne)
        ('BACKGROUND', (0, -2), (-1, -2), WINE_PRIMARY),
        ('TEXTCOLOR', (0, -2), (-1, -2), white),
        ('FONTNAME', (0, -2), (-1, -2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -2), (-1, -2), 8),
        # Ligne légende (dernière ligne) — span sur toutes les colonnes
        ('SPAN', (0, -1), (-1, -1)),
        ('BACKGROUND', (0, -1), (-1, -1), white),
        ('TOPPADDING', (0, -1), (-1, -1), 2*mm),
        # Padding
        ('LEFTPADDING', (0, 0), (-1, -1), 2*mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2*mm),
        ('TOPPADDING', (0, 0), (-1, -2), 1*mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1*mm),
    ]

    # Colorer les recommandations (colonne 12)
    for i, b in enumerate(inv_display, 1):
        color_map = {"À conserver": GREEN_REC, "À vendre": RED_REC, "À surveiller": ORANGE_REC}
        c = color_map.get(b["reco"], TEXT_DARK)
        ts.append(('TEXTCOLOR', (12, i), (12, i), c))
        ts.append(('FONTNAME', (12, i), (12, i), 'Helvetica-Bold'))

    inv_table.setStyle(TableStyle(ts))

    story.append(inv_table)

    # Revenir en portrait pour les sections suivantes
    story.append(NextPageTemplate('content'))

    story.append(PageBreak())

    # ═══ SECTION 3 — ESTIMATION GLOBALE ═══
    story.append(section_number_title("3", "ESTIMATION GLOBALE"))
    story.append(gold_divider())
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "Une fourchette d'estimation avec une marge de +/-15% est présentée afin de tenir compte "
        "des variations de marché, des écarts entre canaux de vente et des limites inhérentes "
        "à l'analyse visuelle lorsqu'elle est réalisée à distance ou sur base documentaire.",
        styles['Intro']))

    # KPI cards
    low = math.ceil(total_val() * 0.85 / 5) * 5
    high = math.ceil(total_val() * 1.10 / 5) * 5

    fourch_str = f'{low:,}€ — {high:,}€'.replace(",", " ")
    kpi_data = [
        [Paragraph(f'{total_btls()}', styles['KPI_Number']),
         Paragraph(f'{len(INVENTAIRE)}', styles['KPI_Number']),
         Paragraph(f'{total_val():,}€'.replace(",", " "), styles['KPI_Number'])],
        [Paragraph('Bouteilles', styles['KPI_Label']),
         Paragraph('Références', styles['KPI_Label']),
         Paragraph('Valeur estimée', styles['KPI_Label'])],
        [Paragraph(
            f'<font size="8" color="{TEXT_MID.hexval()}">Fourchette d\'estimation : </font>'
            f'<font size="10" color="{WINE_PRIMARY.hexval()}"><b>{fourch_str}</b></font>',
            ParagraphStyle('KPIFourch', alignment=TA_CENTER, leading=14)),
         '', ''],
    ]
    kpi_w = (PAGE_W - MARGIN_L - MARGIN_R) / 3
    kpi_table = Table(kpi_data, colWidths=[kpi_w]*3, rowHeights=[10*mm, 5*mm, 8*mm])
    kpi_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('SPAN', (0, 2), (-1, 2)),
        ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
        ('ROUNDEDCORNERS', [3, 3, 3, 3]),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_LIGHT),
    ]))
    story.append(kpi_table)

    # Style commun pour les tableaux de synthèse (5 colonnes)
    def make_synthesis_table(rows_data):
        content_w = PAGE_W - MARGIN_L - MARGIN_R
        tbl = Table(rows_data,
                    colWidths=[content_w * 0.22, content_w * 0.22, content_w * 0.14,
                               content_w * 0.22, content_w * 0.14],
                    rowHeights=[6*mm] * len(rows_data))
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), WINE_DARK),
            ('TEXTCOLOR', (0, 0), (-1, 0), GOLD_LIGHT),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TEXTCOLOR', (0, 1), (-1, -2), TEXT_DARK),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [white, BG_LIGHT]),
            ('LINEBELOW', (0, 0), (-1, -2), 0.2, BORDER_LIGHT),
            ('BACKGROUND', (0, -1), (-1, -1), WINE_PRIMARY),
            ('TEXTCOLOR', (0, -1), (-1, -1), white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3*mm),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3*mm),
            ('TOPPADDING', (0, 0), (-1, -1), 2*mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2*mm),
        ]))
        return tbl

    # Tableau de synthèse par format (si plusieurs formats présents)
    formats_dict = {}
    for b in INVENTAIRE:
        fmt = b.get("format", "75cl")
        if fmt not in formats_dict:
            formats_dict[fmt] = {"qte": 0, "val": 0}
        formats_dict[fmt]["qte"] += b["qte"]
        formats_dict[fmt]["val"] += b["qte"] * b["val_unit"]

    if len(formats_dict) > 1:
        fmt_rows = [["Format", "Nb bouteilles", "Part volume", "Valeur estimée", "Part valeur"]]
        for fmt in sorted(formats_dict.keys()):
            d = formats_dict[fmt]
            pct_vol = f'{d["qte"] / total_btls() * 100:.1f}%'
            pct_val = f'{d["val"] / total_val() * 100:.1f}%'
            fmt_rows.append([fmt, str(d["qte"]), pct_vol,
                             f'{d["val"]:,}€'.replace(",", " "), pct_val])
        fmt_rows.append(["Total", str(total_btls()), "100%",
                         f'{total_val():,}€'.replace(",", " "), "100%"])
        story.append(KeepTogether([
            Spacer(1, 1.5*mm),
            Paragraph("<b>Synthèse par format</b>", styles['SectionSubtitle']),
            make_synthesis_table(fmt_rows),
        ]))

    # Tableau de synthèse par région (si plusieurs régions présentes)
    regions_dict = {}
    for b in INVENTAIRE:
        reg = b["region"]
        if reg not in regions_dict:
            regions_dict[reg] = {"qte": 0, "val": 0}
        regions_dict[reg]["qte"] += b["qte"]
        regions_dict[reg]["val"] += b["qte"] * b["val_unit"]

    if len(regions_dict) > 1:
        reg_rows = [["Région", "Nb bouteilles", "Part volume", "Valeur estimée", "Part valeur"]]
        for reg in sorted(regions_dict.keys()):
            d = regions_dict[reg]
            pct_vol = f'{d["qte"] / total_btls() * 100:.1f}%'
            pct_val = f'{d["val"] / total_val() * 100:.1f}%'
            reg_rows.append([reg, str(d["qte"]), pct_vol,
                             f'{d["val"]:,}€'.replace(",", " "), pct_val])
        reg_rows.append(["Total", str(total_btls()), "100%",
                         f'{total_val():,}€'.replace(",", " "), "100%"])
        story.append(KeepTogether([
            Spacer(1, 1.5*mm),
            Paragraph("<b>Synthèse par région</b>", styles['SectionSubtitle']),
            make_synthesis_table(reg_rows),
        ]))

    # Tableau de synthèse par couleur (si plusieurs couleurs présentes)
    colors_dict = {}
    for b in INVENTAIRE:
        couleur = b["couleur"]
        if couleur not in colors_dict:
            colors_dict[couleur] = {"qte": 0, "val": 0}
        colors_dict[couleur]["qte"] += b["qte"]
        colors_dict[couleur]["val"] += b["qte"] * b["val_unit"]

    if len(colors_dict) > 1:
        col_rows = [["Couleur", "Nb bouteilles", "Part volume", "Valeur estimée", "Part valeur"]]
        for couleur in sorted(colors_dict.keys()):
            d = colors_dict[couleur]
            pct_vol = f'{d["qte"] / total_btls() * 100:.1f}%'
            pct_val = f'{d["val"] / total_val() * 100:.1f}%'
            col_rows.append([couleur, str(d["qte"]), pct_vol,
                             f'{d["val"]:,}€'.replace(",", " "), pct_val])
        col_rows.append(["Total", str(total_btls()), "100%",
                         f'{total_val():,}€'.replace(",", " "), "100%"])
        story.append(KeepTogether([
            Spacer(1, 1.5*mm),
            Paragraph("<b>Synthèse par couleur</b>", styles['SectionSubtitle']),
            make_synthesis_table(col_rows),
        ]))

    story.append(PageBreak())

    # ═══ SECTION 4 — CONTEXTE MARCHÉ ═══
    story.append(section_number_title("4", "CONTEXTE MARCHÉ"))
    story.append(gold_divider())
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "Le marché des vins fins demeure sélectif et sensible à plusieurs facteurs structurants : "
        "qualité de présentation, provenance, désirabilité du domaine, profondeur du millésime, "
        "format et canal de commercialisation. Dans ce contexte, l'estimation d'une cave ne peut "
        "se limiter à une cote théorique ; elle doit être replacée dans une lecture de marché "
        "concrète et actuelle.",
        styles['Intro']))

    cm_content_w = PAGE_W - MARGIN_L - MARGIN_R
    cm_pad = 4*mm

    # ── Bloc Tendances générales du marché ──
    cm_tend_title = Paragraph(
        f'<font color="{WINE_PRIMARY.hexval()}"><b>TENDANCES GÉNÉRALES DU MARCHÉ</b></font>',
        ParagraphStyle('CMTendTitle', fontName='Helvetica-Bold', fontSize=9,
                       spaceBefore=1*mm, spaceAfter=0))
    cm_tend_flowables = [cm_tend_title, Spacer(1, 4*mm)]
    cm_body_style = ParagraphStyle('CMBody', fontName='Helvetica', fontSize=9,
                                    textColor=TEXT_DARK, leading=13)
    for ligne in CLIENT_DATA.get("contexte_marche_tendances", contexte_marche_tendances).split('\n'):
        if ligne.strip():
            cm_tend_flowables.append(Paragraph(ligne.strip(), cm_body_style))
            cm_tend_flowables.append(Spacer(1, 4*mm))
    cm_tend_inner_rows = [[f] for f in cm_tend_flowables]
    cm_tend_inner = Table(cm_tend_inner_rows, colWidths=[cm_content_w - 8*mm])
    cm_tend_inner.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    bloc_tend = Table([[cm_tend_inner]], colWidths=[cm_content_w])
    bloc_tend.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_LIGHT),
        ('LEFTPADDING', (0, 0), (-1, -1), cm_pad),
        ('RIGHTPADDING', (0, 0), (-1, -1), cm_pad),
        ('TOPPADDING', (0, 0), (-1, -1), 3*mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3*mm),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(bloc_tend)
    story.append(Spacer(1, 5*mm))

    # ── Bloc Liquidité des références ──
    cm_liq_title = Paragraph(
        f'<font color="{WINE_PRIMARY.hexval()}"><b>LIQUIDITÉ DES RÉFÉRENCES</b></font>',
        ParagraphStyle('CMLiqTitle', fontName='Helvetica-Bold', fontSize=9,
                       spaceBefore=1*mm, spaceAfter=0))
    cm_liq_flowables = [cm_liq_title, Spacer(1, 4*mm)]
    for ligne in CLIENT_DATA.get("contexte_marche_liquidite", contexte_marche_liquidite).split('\n'):
        if ligne.strip():
            cm_liq_flowables.append(Paragraph(ligne.strip(), cm_body_style))
            cm_liq_flowables.append(Spacer(1, 4*mm))
    cm_liq_inner_rows = [[f] for f in cm_liq_flowables]
    cm_liq_inner = Table(cm_liq_inner_rows, colWidths=[cm_content_w - 8*mm])
    cm_liq_inner.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    bloc_liq = Table([[cm_liq_inner]], colWidths=[cm_content_w])
    bloc_liq.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_LIGHT),
        ('LEFTPADDING', (0, 0), (-1, -1), cm_pad),
        ('RIGHTPADDING', (0, 0), (-1, -1), cm_pad),
        ('TOPPADDING', (0, 0), (-1, -1), 3*mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3*mm),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(bloc_liq)

    story.append(PageBreak())

    # ═══ SECTION 5 — GRAPHIQUES DE RÉPARTITION — PAGE 1 ═══
    story.append(section_number_title("5", "GRAPHIQUES DE RÉPARTITION"))
    story.append(gold_divider())
    story.append(Spacer(1, 2*mm))

    full_w = PAGE_W - MARGIN_L - MARGIN_R
    chart_w_60 = full_w * 0.60
    chart_h_pie = chart_w_60 * 0.72
    chart_h_bar = chart_w_60 * 0.50
    sp_chart = Spacer(1, 20*mm)

    # Génération des graphiques
    chart_region = create_region_chart()
    chart_color = create_color_chart()
    chart_reco = create_recommendation_chart()
    chart_apogee = create_apogee_chart()
    chart_region_val = create_region_value_chart()

    # Auto-generated region analysis
    reg_vol = {}
    reg_val = {}
    for b in INVENTAIRE:
        reg_vol[b["region"]] = reg_vol.get(b["region"], 0) + b["qte"]
        reg_val[b["region"]] = reg_val.get(b["region"], 0) + b["qte"] * b["val_unit"]
    top_vol_reg = max(reg_vol, key=reg_vol.get) if reg_vol else "N/A"
    top_vol_pct = int(reg_vol.get(top_vol_reg, 0) / total_btls() * 100) if total_btls() > 0 else 0
    nb_regions = len(reg_vol)
    top_val_reg = max(reg_val, key=reg_val.get) if reg_val else "N/A"
    top_val_pct = int(reg_val.get(top_val_reg, 0) / total_val() * 100) if total_val() > 0 else 0
    top2_val = sorted(reg_val.items(), key=lambda x: x[1], reverse=True)[:2]
    top2_val_pct = int(sum(v for _, v in top2_val) / total_val() * 100) if total_val() > 0 else 0

    # ── Graphique 1 : Répartition bouteilles par région ──
    story.append(Paragraph("<b>Répartition des bouteilles par région</b>", styles['SectionSubtitle']))
    story.append(Paragraph(
        f"La cave couvre {nb_regions} régions viticoles distinctes. Le {top_vol_reg} domine en volume "
        f"avec {top_vol_pct}% des bouteilles, ce qui confère à la cave une identité régionale "
        f"{'marquée' if top_vol_pct >= 40 else 'diversifiée'}.",
        styles['Intro']))
    img_region = safe_chart_image(chart_region, width=chart_w_60, height=chart_h_pie)
    img_region_wrapper = Table([[img_region]], colWidths=[full_w])
    img_region_wrapper.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER')]))
    story.append(img_region_wrapper)
    story.append(sp_chart)

    # ── Graphique 2 : Valeur par région ──
    story.append(Paragraph("<b>Répartition de la valeur par région (€)</b>", styles['SectionSubtitle']))
    if len(top2_val) >= 2:
        region_val_text = (
            f"En valeur, le {top_val_reg} concentre {top_val_pct}% de l'estimation totale. "
            f"Les deux premières régions ({top2_val[0][0]} et {top2_val[1][0]}) représentent "
            f"à elles seules {top2_val_pct}% de la valeur patrimoniale de la cave.")
    elif len(top2_val) == 1:
        region_val_text = (
            f"En valeur, le {top_val_reg} concentre {top_val_pct}% de l'estimation totale, "
            f"constituant le pôle de valeur unique de cette cave.")
    else:
        region_val_text = "Données insuffisantes pour analyser la répartition de valeur par région."
    story.append(Paragraph(region_val_text, styles['Intro']))
    img_region_val = safe_chart_image(chart_region_val, width=chart_w_60, height=chart_h_pie)
    img_region_val_wrapper = Table([[img_region_val]], colWidths=[full_w])
    img_region_val_wrapper.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER')]))
    story.append(img_region_val_wrapper)

    story.append(PageBreak())

    # ═══ SECTION 5 — PAGE 2 ═══
    story.append(section_number_title("5", "GRAPHIQUES DE RÉPARTITION (suite)"))
    story.append(gold_divider())
    story.append(Spacer(1, 2*mm))

    # ── Graphique 3 : Répartition par couleur ──
    color_stats = {}
    for b in INVENTAIRE:
        c = b["couleur"]
        if c not in color_stats:
            color_stats[c] = {"qte": 0, "val": 0}
        color_stats[c]["qte"] += b["qte"]
        color_stats[c]["val"] += b["qte"] * b["val_unit"]

    dominant_color = max(color_stats, key=lambda c: color_stats[c]["qte"]) if color_stats else "Rouge"
    dc_vol_pct = int(color_stats[dominant_color]["qte"] / total_btls() * 100) if total_btls() > 0 else 0
    dc_val_pct = int(color_stats[dominant_color]["val"] / total_val() * 100) if total_val() > 0 else 0

    story.append(Paragraph("<b>Répartition par couleur</b>", styles['SectionSubtitle']))
    story.append(Paragraph(
        f"Les vins {dominant_color.lower()}s représentent {dc_vol_pct}% du volume et "
        f"{dc_val_pct}% de la valeur estimée, constituant le cœur patrimonial de cette cave.",
        styles['Intro']))
    img_color = safe_chart_image(chart_color, width=chart_w_60, height=chart_h_pie)
    img_color_wrapper = Table([[img_color]], colWidths=[full_w])
    img_color_wrapper.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER')]))
    story.append(img_color_wrapper)
    story.append(sp_chart)

    # ── Graphique 4 : Potentiel de garde ──
    current_year = 2026
    dev_btls = 0
    for b in INVENTAIRE:
        apogee = b["apogee"]
        if apogee != "À boire":
            try:
                parts = apogee.replace("–", "-").split("-")
                start = int(parts[0])
                end = int(parts[1])
                mid = (start + end) / 2
                if mid - current_year > 3:
                    dev_btls += b["qte"]
            except:
                pass

    story.append(Paragraph("<b>Potentiel de garde</b>", styles['SectionSubtitle']))
    story.append(Paragraph(
        f"{dev_btls} bouteilles sur {total_btls()} sont encore en phase de développement, "
        f"suggérant un potentiel de valorisation à moyen terme.",
        styles['Intro']))
    img_apogee = safe_chart_image(chart_apogee, width=chart_w_60, height=chart_h_bar)
    img_apogee_wrapper = Table([[img_apogee]], colWidths=[full_w])
    img_apogee_wrapper.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER')]))
    story.append(img_apogee_wrapper)

    story.append(PageBreak())

    # ═══ SECTION 5 — PAGE 3 : Recommandations + Top 5 ═══
    story.append(section_number_title("5", "GRAPHIQUES DE RÉPARTITION (suite)"))
    story.append(gold_divider())
    story.append(Spacer(1, 2*mm))

    # ── Graphique 5 : Recommandations ──
    reco_g = {"À conserver": {"btls": 0, "val": 0}, "À vendre": {"btls": 0, "val": 0}, "À surveiller": {"btls": 0, "val": 0}}
    for b in INVENTAIRE:
        reco_g[b["reco"]]["btls"] += b["qte"]
        reco_g[b["reco"]]["val"] += b["qte"] * b["val_unit"]
    vendre_pct = int(reco_g["À vendre"]["val"] / total_val() * 100) if total_val() > 0 else 0
    garder_pct = int(reco_g["À conserver"]["val"] / total_val() * 100) if total_val() > 0 else 0

    story.append(Paragraph("<b>Recommandations</b>", styles['SectionSubtitle']))
    story.append(Paragraph(
        f"{vendre_pct}% de la valeur estimée est orientée vers la vente, "
        f"{garder_pct}% vers la conservation. L'arbitrage proposé reflète "
        f"l'équilibre entre potentiel de garde et opportunité de marché.",
        styles['Intro']))
    img_reco = safe_chart_image(chart_reco, width=chart_w_60, height=chart_h_bar)
    img_reco_wrapper = Table([[img_reco]], colWidths=[full_w])
    img_reco_wrapper.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER')]))
    story.append(img_reco_wrapper)
    story.append(Spacer(1, 10*mm))

    # ── Concentration des valeurs ──
    story.append(Paragraph("<b>Concentration des valeurs</b>", styles['SectionSubtitle']))
    story.append(Paragraph(
        "Les références ci-dessous concentrent une part significative de la valeur totale de la cave. "
        "Elles constituent les principaux points d'attention, tant pour la conservation que pour "
        "une éventuelle cession. Toute décision relative à ces références doit donc être prise "
        "avec un niveau d'attention supérieur.",
        styles['Intro']))
    story.append(Spacer(1, 1*mm))

    # ── Top 5 bouteilles — valeur unitaire ──
    sorted_unit = sorted(INVENTAIRE, key=lambda x: x["val_unit"], reverse=True)[:5]
    story.append(Paragraph("<b>Top 5 bouteilles — valeur unitaire</b>", styles['SectionSubtitle']))

    top_unit_rows = [["Domaine / Cuvée", "Appellation", "Couleur", "Format", "Mill.", "Val. unit.", "Orientation"]]
    top5_cell_style = ParagraphStyle('Top5Cell', fontName='Helvetica', fontSize=7,
                                      textColor=TEXT_DARK, leading=9)
    for b in sorted_unit:
        mill_str = str(b["millesime"]) if b["millesime"] != 0 else "NM"
        top_unit_rows.append([
            Paragraph(b["bouteille"], top5_cell_style),
            Paragraph(b["appellation"], top5_cell_style),
            b["couleur"],
            b.get("format", "75cl"), mill_str,
            f'{b["val_unit"]:,}€'.replace(",", " "), b["reco"]])

    top_unit_tbl = Table(top_unit_rows,
                         colWidths=[full_w*0.26, full_w*0.26, full_w*0.09,
                                    full_w*0.07, full_w*0.07, full_w*0.13, full_w*0.12])
    top_unit_ts = [
        ('BACKGROUND', (0, 0), (-1, 0), WINE_DARK),
        ('TEXTCOLOR', (0, 0), (-1, 0), GOLD_LIGHT),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TEXTCOLOR', (0, 1), (-1, -1), TEXT_DARK),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (2, 1), (4, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, BG_LIGHT]),
        ('LINEBELOW', (0, 0), (-1, -1), 0.2, BORDER_LIGHT),
        ('LEFTPADDING', (0, 0), (-1, -1), 2*mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2*mm),
    ]
    for i, b in enumerate(sorted_unit, 1):
        color_map = {"À conserver": GREEN_REC, "À vendre": RED_REC, "À surveiller": ORANGE_REC}
        c = color_map.get(b["reco"], TEXT_DARK)
        top_unit_ts.append(('TEXTCOLOR', (6, i), (6, i), c))
        top_unit_ts.append(('FONTNAME', (6, i), (6, i), 'Helvetica-Bold'))
    top_unit_tbl.setStyle(TableStyle(top_unit_ts))
    story.append(top_unit_tbl)
    story.append(Spacer(1, 4*mm))

    # ── Top 5 lots — valeur totale (pleine largeur, avec appellation, couleur, format) ──
    sorted_total = sorted(INVENTAIRE, key=lambda x: x["qte"] * x["val_unit"], reverse=True)[:5]
    story.append(Paragraph("<b>Top 5 lots — valeur totale</b>", styles['SectionSubtitle']))

    top_total_rows = [["Domaine / Cuvée", "Appellation", "Couleur", "Format", "Mill.", "Qté", "Val. totale", "Orientation"]]
    for b in sorted_total:
        mill_str = str(b["millesime"]) if b["millesime"] != 0 else "NM"
        val_t = b["qte"] * b["val_unit"]
        top_total_rows.append([
            Paragraph(b["bouteille"], top5_cell_style),
            Paragraph(b["appellation"], top5_cell_style),
            b["couleur"],
            b.get("format", "75cl"), mill_str, str(b["qte"]),
            f'{val_t:,}€'.replace(",", " "), b["reco"]])

    top_total_tbl = Table(top_total_rows,
                          colWidths=[full_w*0.26, full_w*0.26, full_w*0.09,
                                     full_w*0.07, full_w*0.07, full_w*0.04, full_w*0.09, full_w*0.12])
    top_total_ts = [
        ('BACKGROUND', (0, 0), (-1, 0), WINE_DARK),
        ('TEXTCOLOR', (0, 0), (-1, 0), GOLD_LIGHT),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TEXTCOLOR', (0, 1), (-1, -1), TEXT_DARK),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (2, 1), (5, -1), 'CENTER'),
        ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
        ('ALIGN', (7, 1), (7, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, BG_LIGHT]),
        ('LINEBELOW', (0, 0), (-1, -1), 0.2, BORDER_LIGHT),
        ('LEFTPADDING', (0, 0), (-1, -1), 2*mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2*mm),
    ]
    for i, b in enumerate(sorted_total, 1):
        color_map = {"À conserver": GREEN_REC, "À vendre": RED_REC, "À surveiller": ORANGE_REC}
        c = color_map.get(b["reco"], TEXT_DARK)
        top_total_ts.append(('TEXTCOLOR', (7, i), (7, i), c))
        top_total_ts.append(('FONTNAME', (7, i), (7, i), 'Helvetica-Bold'))
    top_total_tbl.setStyle(TableStyle(top_total_ts))
    story.append(top_total_tbl)

    story.append(PageBreak())

    # ═══ SECTION 6 — RECOMMANDATIONS ═══
    story.append(section_number_title("6", "RECOMMANDATIONS"))
    story.append(gold_divider())
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "Pour chaque lot, une recommandation claire est formulée : à conserver, à vendre ou à surveiller. "
        "Chaque préconisation est justifiée par l'état de la bouteille, sa cote actuelle, "
        "son potentiel de garde et les tendances de marché.",
        styles['Intro']))

    # Regrouper par recommandation
    groups = {"À conserver": [], "À vendre": [], "À surveiller": []}
    for b in INVENTAIRE:
        groups[b["reco"]].append(b)

    # Phrases contextuelles par code de justification
    # (codes définis dans JUSTIFICATION_CODES)

    reco_config = [
        ("À conserver", GREEN_REC, CLIENT_DATA.get("texte_intro_garder",
         "Sont classés dans cette catégorie les lots présentant un potentiel d'évolution, "
         "une cohérence patrimoniale ou une meilleure perspective d'arbitrage à moyen terme.")),
        ("À vendre", RED_REC, CLIENT_DATA.get("texte_intro_vendre",
         "Sont classés dans cette catégorie les lots dont la fenêtre de marché paraît favorable, "
         "dont la liquidité semble satisfaisante ou dont l'intérêt principal réside aujourd'hui "
         "dans une valorisation patrimoniale ou commerciale à court terme.")),
        ("À surveiller", ORANGE_REC, CLIENT_DATA.get("texte_intro_attendre",
         "Sont classés dans cette catégorie les lots pour lesquels une décision immédiate apparaît "
         "moins pertinente, soit en raison de leur maturité, soit en raison d'un manque de visibilité "
         "suffisant sur leur fenêtre optimale d'arbitrage.")),
    ]

    reco_content_w = PAGE_W - MARGIN_L - MARGIN_R

    for reco_name, reco_color, reco_intro in reco_config:
        bottles = groups.get(reco_name, [])
        if not bottles:
            continue

        total_reco = sum(b["qte"] * b["val_unit"] for b in bottles)
        reco_btls = sum(b["qte"] for b in bottles)

        # Collecter les éléments du bloc pour KeepTogether
        block_elements = []

        # Titre de recommandation
        block_elements.append(Paragraph(
            f'<font color="{reco_color.hexval()}"><b>▎ {reco_name.upper()}</b></font>'
            f'&nbsp;&nbsp;<font color="{TEXT_LIGHT.hexval()}" size="8">'
            f'{reco_btls} bouteilles · {total_reco:,}€</font>'.replace(",", " "),
            ParagraphStyle('RecoTitle', fontName='Helvetica-Bold', fontSize=12,
                           leading=16, spaceBefore=5*mm, spaceAfter=2*mm)))

        block_elements.append(Paragraph(f'<i>{reco_intro}</i>',
                               ParagraphStyle('RecoIntro', fontName='Helvetica-Oblique',
                                              fontSize=8.5, textColor=TEXT_MID, leading=12,
                                              spaceAfter=3*mm, alignment=TA_JUSTIFY)))

        # Tableau — 2 lignes par lot : données factuelles + analyse
        is_conserver = (reco_name == "À conserver")

        if is_conserver:
            header = ["Domaine / Cuvée", "Appellation", "Couleur", "Mill.", "Format", "Qté", "Val. totale", "Durée de garde"]
            col_widths = [reco_content_w * 0.20, reco_content_w * 0.21,
                          reco_content_w * 0.08, reco_content_w * 0.06,
                          reco_content_w * 0.07, reco_content_w * 0.05, reco_content_w * 0.10, reco_content_w * 0.11]
        else:
            header = ["Domaine / Cuvée", "Appellation", "Couleur", "Mill.", "Format", "Qté", "Val. totale"]
            col_widths = [reco_content_w * 0.24, reco_content_w * 0.25,
                          reco_content_w * 0.08, reco_content_w * 0.06,
                          reco_content_w * 0.07, reco_content_w * 0.05, reco_content_w * 0.13]

        nb_cols = len(header)
        reco_rows = [header]

        domaine_style = ParagraphStyle('DomaineCell', fontName='Helvetica', fontSize=7,
                                        textColor=TEXT_DARK, leading=9)
        analyse_style = ParagraphStyle('AnalyseCell', fontName='Helvetica-BoldOblique', fontSize=7,
                                        textColor=TEXT_MID, leading=9)

        for b in bottles:
            mill_str = str(b["millesime"]) if b["millesime"] != 0 else "NM"
            val_t = b["qte"] * b["val_unit"]
            code = b.get("justification_code", None)
            justif = JUSTIFICATION_CODES.get(code, "[À compléter]") if code else "[À compléter]"
            note = b.get("note_marche", "")

            # Ligne 1 — données factuelles
            row1 = [
                Paragraph(b["bouteille"], domaine_style),
                Paragraph(b["appellation"], domaine_style),
                b["couleur"],
                mill_str,
                b.get("format", "75cl"),
                str(b["qte"]),
                f'{val_t:,}€'.replace(",", " "),
            ]
            if is_conserver:
                dg = b.get("duree_garde", "")
                row1.append(dg if dg else "—")
            reco_rows.append(row1)

            # Ligne 2 — analyse et justification (spanning)
            analyse_text = justif
            if note:
                analyse_text += f" · {note}"
            row2 = [Paragraph(analyse_text, analyse_style)] + [''] * (nb_cols - 1)
            reco_rows.append(row2)

        reco_table = Table(reco_rows, colWidths=col_widths, repeatRows=1)

        # Build style
        reco_ts = [
            ('BACKGROUND', (0, 0), (-1, 0), reco_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TEXTCOLOR', (0, 1), (-1, -1), TEXT_DARK),
            ('ALIGN', (2, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),   # Couleur
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),   # Mill.
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),   # Format
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),   # Qté
            ('ALIGN', (6, 1), (6, -1), 'RIGHT'),    # Val. totale
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 2*mm),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2*mm),
            ('TOPPADDING', (0, 0), (-1, -1), 1.5*mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5*mm),
        ]

        if is_conserver:
            reco_ts.append(('ALIGN', (7, 1), (7, -1), 'CENTER'))

        # Style par paire de lignes (ligne 1 + ligne 2 = un lot)
        for idx in range(len(bottles)):
            row1_idx = 1 + idx * 2   # Ligne données
            row2_idx = 2 + idx * 2   # Ligne analyse

            # SPAN la ligne 2 sur toutes les colonnes
            reco_ts.append(('SPAN', (0, row2_idx), (-1, row2_idx)))

            # Alternance de fond par lot (pas par ligne)
            bg = white if idx % 2 == 0 else BG_LIGHT
            reco_ts.append(('BACKGROUND', (0, row1_idx), (-1, row1_idx), bg))
            reco_ts.append(('BACKGROUND', (0, row2_idx), (-1, row2_idx), bg))

            # Pas de ligne entre row1 et row2 du même lot
            reco_ts.append(('TOPPADDING', (0, row2_idx), (-1, row2_idx), 0))
            reco_ts.append(('BOTTOMPADDING', (0, row1_idx), (-1, row1_idx), 0.5*mm))

            # Séparateur entre lots
            if idx < len(bottles) - 1:
                reco_ts.append(('LINEBELOW', (0, row2_idx), (-1, row2_idx), 0.2, BORDER_LIGHT))

        reco_table.setStyle(TableStyle(reco_ts))

        # Garder titre + intro ensemble (évite un titre orphelin en bas de page),
        # mais laisser la table couler naturellement — sinon un gros bloc À conserver
        # pousse toute une page blanche devant lui.
        story.append(KeepTogether(block_elements))
        story.append(reco_table)
        story.append(Spacer(1, 3*mm))

    # ═══ SECTION 7 — POTENTIEL DE GARDE ═══
    story.append(PageBreak())
    story.append(section_number_title("7", "POTENTIEL DE GARDE"))
    story.append(gold_divider())
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "Pour chaque lot, l'apogée estimée indique la fenêtre optimale de <b>consommation</b> "
        "— la période pendant laquelle le vin exprimera pleinement son potentiel gustatif. "
        "Ces estimations sont établies à titre indicatif sur la base du millésime, de l'appellation "
        "et du style du domaine.",
        styles['Intro']))

    # Classer les bouteilles par catégorie de garde
    current_year = 2026
    garde_cats = {
        "À boire maintenant": [],
        "Proche apogée (1–3 ans)": [],
        "En développement (3–10 ans)": [],
        "Garde longue (10 ans et plus)": [],
    }

    for b in INVENTAIRE:
        apogee = b.get("apogee", "")
        if b["millesime"] == 0 or apogee in ("À boire maintenant", "À boire", ""):
            garde_cats["À boire maintenant"].append(b)
        else:
            try:
                parts = apogee.replace("–", "-").split("-")
                start = int(parts[0])
                end = int(parts[1])
                mid = (start + end) / 2
                diff = mid - current_year
                if diff <= 0:
                    garde_cats["À boire maintenant"].append(b)
                elif diff <= 3:
                    garde_cats["Proche apogée (1–3 ans)"].append(b)
                elif diff <= 10:
                    garde_cats["En développement (3–10 ans)"].append(b)
                else:
                    garde_cats["Garde longue (10 ans et plus)"].append(b)
            except (ValueError, IndexError):
                garde_cats["À boire maintenant"].append(b)

    garde_config = [
        ("À boire maintenant", HexColor("#B33A3A"), "Bouteilles à apogée ou au-delà"),
        ("Proche apogée (1–3 ans)", HexColor("#D4841A"), "Fenêtre qui s'ouvre bientôt"),
        ("En développement (3–10 ans)", HexColor("#C5A258"), "Patience requise"),
        ("Garde longue (10 ans et plus)", HexColor("#2E7D4F"), "Patrimoine à long terme"),
    ]

    garde_content_w = PAGE_W - MARGIN_L - MARGIN_R
    garde_domaine_style = ParagraphStyle('GardeDomaine', fontName='Helvetica', fontSize=7,
                                          textColor=TEXT_DARK, leading=9)

    for cat_name, cat_color, cat_desc in garde_config:
        bottles = garde_cats.get(cat_name, [])
        if not bottles:
            continue

        # Tri par millésime croissant, NM à la fin
        bottles_sorted = sorted(bottles, key=lambda b: (
            0 if b["millesime"] != 0 else 1,
            b["millesime"] if b["millesime"] != 0 else 9999,
            b["bouteille"].lower()
        ))

        total_btls_cat = sum(b["qte"] for b in bottles_sorted)

        story.append(Paragraph(
            f'<font color="{cat_color.hexval()}"><b>▎ {cat_name.upper()}</b></font>'
            f'&nbsp;&nbsp;<font color="{TEXT_LIGHT.hexval()}" size="8">'
            f'{total_btls_cat} bouteilles — {cat_desc}</font>',
            ParagraphStyle('GardeTitle', fontName='Helvetica-Bold', fontSize=11,
                           leading=15, spaceBefore=5*mm, spaceAfter=2*mm)))

        garde_rows = [["Domaine / Cuvée", "Appellation", "Couleur", "Mill.", "Qté", "Apogée estimée", "Orientation"]]
        for b in bottles_sorted:
            mill_str = str(b["millesime"]) if b["millesime"] != 0 else "NM"
            garde_rows.append([
                Paragraph(b["bouteille"], garde_domaine_style),
                Paragraph(b["appellation"], garde_domaine_style),
                b["couleur"],
                mill_str,
                str(b["qte"]),
                b.get("apogee", "—"),
                b["reco"],
            ])

        garde_tbl = Table(garde_rows,
                          colWidths=[garde_content_w * 0.22, garde_content_w * 0.22,
                                     garde_content_w * 0.08, garde_content_w * 0.06,
                                     garde_content_w * 0.05, garde_content_w * 0.17, garde_content_w * 0.20],
                          repeatRows=1)
        garde_ts = [
            ('BACKGROUND', (0, 0), (-1, 0), cat_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TEXTCOLOR', (0, 1), (-1, -1), TEXT_DARK),
            ('ALIGN', (2, 0), (5, -1), 'CENTER'),
            ('ALIGN', (6, 0), (6, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, BG_LIGHT]),
            ('LINEBELOW', (0, 0), (-1, -1), 0.2, BORDER_LIGHT),
            ('LEFTPADDING', (0, 0), (-1, -1), 2*mm),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2*mm),
            ('TOPPADDING', (0, 0), (-1, -1), 1.5*mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5*mm),
        ]
        # Colorer les recommandations
        for i, b in enumerate(bottles_sorted, 1):
            color_map = {"À conserver": GREEN_REC, "À vendre": RED_REC, "À surveiller": ORANGE_REC}
            c = color_map.get(b["reco"], TEXT_DARK)
            garde_ts.append(('TEXTCOLOR', (6, i), (6, i), c))
            garde_ts.append(('FONTNAME', (6, i), (6, i), 'Helvetica-Bold'))
        garde_tbl.setStyle(TableStyle(garde_ts))
        story.append(garde_tbl)
        story.append(Spacer(1, 3*mm))

    # ═══ PAGE 1 — PLAN D'ACTION (personnalisée) ═══
    story.append(PageBreak())

    story.append(section_number_title("8", "PLAN D'ACTION"))
    story.append(gold_divider())
    story.append(Spacer(1, 4*mm))

    pa_content_w = PAGE_W - MARGIN_L - MARGIN_R

    # ── Bloc Conclusion de l'expert ──
    aide = CLIENT_DATA.get("aide_decision", "")
    aide_default = "Ex : Cave bien construite"
    if aide and not aide.startswith("Ex :"):
        bloc_aide_title = Paragraph(
            '<font color="#C9A84C"><b>CONCLUSION DE L\'EXPERT</b></font>',
            ParagraphStyle('AideTitle', fontName='Helvetica-Bold', fontSize=9,
                           spaceBefore=1*mm, spaceAfter=0))
        aide_style = ParagraphStyle('AideBody', fontName='Helvetica', fontSize=9,
                                     textColor=HexColor("#FAF6F0"), leading=13)
        aide_flowables = [bloc_aide_title, Spacer(1, 4*mm)]
        for ligne in aide.split('\n'):
            if ligne.strip():
                aide_flowables.append(Paragraph(ligne.strip(), aide_style))
                aide_flowables.append(Spacer(1, 4*mm))
        aide_inner_rows = [[f] for f in aide_flowables]
        aide_inner = Table(aide_inner_rows, colWidths=[pa_content_w - 14*mm])
        aide_inner.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        bloc_aide = Table([[aide_inner]], colWidths=[pa_content_w])
        bloc_aide.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HexColor("#3A2035")),
            ('BOX', (0, 0), (-1, -1), 0.8, GOLD),
            ('ROUNDEDCORNERS', [3, 3, 3, 3]),
            ('LEFTPADDING', (0, 0), (-1, -1), 4*mm),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4*mm),
            ('TOPPADDING', (0, 0), (-1, -1), 4*mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4*mm),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(bloc_aide)
        story.append(Spacer(1, 3*mm))

    # ── Bloc Points de vigilance ──
    vigilance = CLIENT_DATA.get("points_vigilance", "")
    if vigilance and not vigilance.startswith("Ex :"):
        bloc_vig_title = Paragraph(
            '<font color="#C9A84C"><b>POINTS DE VIGILANCE</b></font>',
            ParagraphStyle('VigTitle', fontName='Helvetica-Bold', fontSize=9,
                           spaceBefore=1*mm, spaceAfter=0))
        vig_style = ParagraphStyle('VigBody', fontName='Helvetica', fontSize=9,
                                    textColor=HexColor("#FAF6F0"), leading=13)
        vig_flowables = [bloc_vig_title, Spacer(1, 4*mm)]
        for ligne in vigilance.split('\n'):
            if ligne.strip():
                vig_flowables.append(Paragraph(ligne.strip(), vig_style))
                vig_flowables.append(Spacer(1, 4*mm))
        vig_inner_rows = [[f] for f in vig_flowables]
        vig_inner = Table(vig_inner_rows, colWidths=[pa_content_w - 14*mm])
        vig_inner.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        bloc_vig = Table([[vig_inner]], colWidths=[pa_content_w])
        bloc_vig.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HexColor("#3A2035")),
            ('BOX', (0, 0), (-1, -1), 0.8, GOLD),
            ('ROUNDEDCORNERS', [3, 3, 3, 3]),
            ('LEFTPADDING', (0, 0), (-1, -1), 4*mm),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4*mm),
            ('TOPPADDING', (0, 0), (-1, -1), 4*mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4*mm),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(bloc_vig)
        story.append(Spacer(1, 3*mm))

    # Grand encadré bordeaux foncé avec étapes
    # Structure attendue : chaque étape est séparée par \n\n (paragraphe).
    # Au sein d'une étape, la 1re ligne = titre (puce or), les lignes suivantes = corps (sans puce, indenté).
    etapes_raw = CLIENT_DATA.get("prochaines_etapes", "")
    etapes_blocks = [blk.strip() for blk in etapes_raw.strip().split("\n\n") if blk.strip()]

    # Construire les lignes d'étapes avec puces or
    etape_flowables = []
    # Sous-titre en or — même format que Conclusion de l'expert / Points de vigilance
    etape_flowables.append(Paragraph(
        '<font color="#C9A84C"><b>VOS PROCHAINES ÉTAPES</b></font>',
        ParagraphStyle('EtapeSubtitle', fontName='Helvetica-Bold', fontSize=9,
                       spaceBefore=1*mm, spaceAfter=0)))
    etape_flowables.append(Spacer(1, 4*mm))

    etape_title_style = ParagraphStyle('EtapeTitle', fontName='Helvetica', fontSize=9.5,
                                        textColor=HexColor("#FAF6F0"), leading=14)
    etape_body_style = ParagraphStyle('EtapeBody', fontName='Helvetica', fontSize=9.5,
                                       textColor=HexColor("#FAF6F0"), leading=14,
                                       leftIndent=5*mm)

    for block in etapes_blocks:
        lines = [l.strip() for l in block.split("\n") if l.strip()]
        if not lines:
            continue
        # 1re ligne : titre avec puce or
        title = lines[0]
        title = title.lstrip("0123456789.").strip() if title[:1].isdigit() else title
        etape_flowables.append(Paragraph(
            f'<font color="#C9A84C"><b>•</b></font>&nbsp;&nbsp;{title}',
            etape_title_style))
        # Lignes suivantes : corps sans puce, indenté
        for body_line in lines[1:]:
            etape_flowables.append(Paragraph(body_line, etape_body_style))
        etape_flowables.append(Spacer(1, 4*mm))

    # Wrapper bordeaux foncé
    inner_rows = [[f] for f in etape_flowables]
    inner_tbl = Table(inner_rows, colWidths=[pa_content_w - 14*mm])
    inner_tbl.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    outer_tbl = Table([[inner_tbl]], colWidths=[pa_content_w])
    outer_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), WINE_DARK),
        ('BOX', (0, 0), (-1, -1), 1.2, GOLD),
        ('ROUNDEDCORNERS', [4, 4, 4, 4]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4*mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4*mm),
        ('TOPPADDING', (0, 0), (-1, -1), 4*mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4*mm),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(outer_tbl)

    # ═══ PAGE 2 — CLÔTURE (générique) ═══
    story.append(PageBreak())

    story.append(section_number_title("8", "PLAN D'ACTION (suite)"))
    story.append(gold_divider())
    story.append(Spacer(1, 4*mm))

    pa_pad = 4*mm
    contact_email = CLIENT_DATA.get("contact_email", "contact@estimationcave.com")

    # ── Bloc Missions complémentaires disponibles ──
    pa_b2_title = Paragraph(
        f'<font color="{WINE_PRIMARY.hexval()}"><b>MISSIONS COMPLÉMENTAIRES DISPONIBLES</b></font>',
        ParagraphStyle('PAB2T', fontName='Helvetica-Bold', fontSize=9,
                       spaceBefore=1*mm, spaceAfter=2*mm))
    pa_b2_body = Paragraph(
        "Ce rapport peut être complété par des missions d'accompagnement sur mesure, "
        "selon vos besoins et l'évolution de votre projet :<br/><br/>"
        "— <b>Inspection physique des flacons</b> : examen approfondi en cave, avec déplacement "
        "ou recours à un mandataire local. Permet de confirmer l'état réel des bouteilles "
        "et d'affiner l'estimation.<br/>"
        "— <b>Accompagnement à la mise en vente</b> : sélection de la maison de vente ou "
        "plateforme adaptée, constitution du dossier de présentation des lots, coordination "
        "avec les experts.<br/>"
        "— <b>Suivi des adjudications</b> : présence ou suivi à distance lors de la vente, "
        "compte-rendu détaillé des résultats et analyse des écarts avec l'estimation initiale.<br/>"
        "— <b>Réévaluation périodique</b> : mise à jour annuelle ou semestrielle de l'estimation, "
        "intégrant les évolutions du marché et l'état de conservation de la cave.<br/>"
        "— <b>Conseil à l'achat</b> : identification de références à acquérir pour compléter, "
        "rééquilibrer ou valoriser la cave, selon un budget et des objectifs définis.<br/><br/>"
        f"Ces missions font l'objet d'un devis personnalisé. Pour toute demande : {contact_email}",
        styles['BodyText2'])

    bloc_pa2 = Table([[pa_b2_title], [pa_b2_body]], colWidths=[pa_content_w])
    bloc_pa2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_LIGHT),
        ('LINEBELOW', (0, 0), (0, 0), 0.8, GOLD),
        ('LEFTPADDING', (0, 0), (-1, -1), pa_pad),
        ('RIGHTPADDING', (0, 0), (-1, -1), pa_pad),
        ('TOPPADDING', (0, 0), (-1, -1), 2*mm),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 3*mm),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(bloc_pa2)
    story.append(Spacer(1, 8*mm))

    # ── Bloc Confidentialité réciproque ──
    pa_b3_title = Paragraph(
        f'<font color="{WINE_PRIMARY.hexval()}"><b>CONFIDENTIALITÉ RÉCIPROQUE</b></font>',
        ParagraphStyle('PAB3T', fontName='Helvetica-Bold', fontSize=9,
                       spaceBefore=1*mm, spaceAfter=2*mm))
    pa_b3_body = Paragraph(
        "Ce rapport est établi dans le cadre d'une relation de stricte confidentialité. "
        "Le client s'engage à ne pas diffuser ce document à des tiers non autorisés. "
        "L'expert s'engage en retour à ne divulguer aucune information relative à l'existence, "
        "la composition, la localisation ou la valeur de la cave à quelque tiers que ce soit, "
        "sans accord écrit préalable du client.",
        styles['BodyText2'])

    bloc_pa3 = Table([[pa_b3_title], [pa_b3_body]], colWidths=[pa_content_w])
    bloc_pa3.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_LIGHT),
        ('LINEBELOW', (0, 0), (0, 0), 0.8, GOLD),
        ('LEFTPADDING', (0, 0), (-1, -1), pa_pad),
        ('RIGHTPADDING', (0, 0), (-1, -1), pa_pad),
        ('TOPPADDING', (0, 0), (-1, -1), 2*mm),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 3*mm),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(bloc_pa3)
    story.append(Spacer(1, 8*mm))

    # ── Bloc Validité et contact ──
    pa_b4_title = Paragraph(
        f'<font color="{WINE_PRIMARY.hexval()}"><b>VALIDITÉ ET CONTACT</b></font>',
        ParagraphStyle('PAB4T', fontName='Helvetica-Bold', fontSize=9,
                       spaceBefore=1*mm, spaceAfter=2*mm))
    pa_b4_body = Paragraph(
        "Les estimations contenues dans ce rapport sont valables 30 jours à compter "
        f"de la date d'émission. Au-delà, une réévaluation peut être demandée. "
        f"Pour toute question : {contact_email}",
        styles['BodyText2'])

    bloc_pa4 = Table([[pa_b4_title], [pa_b4_body]], colWidths=[pa_content_w])
    bloc_pa4.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_LIGHT),
        ('LINEBELOW', (0, 0), (0, 0), 0.8, GOLD),
        ('LEFTPADDING', (0, 0), (-1, -1), pa_pad),
        ('RIGHTPADDING', (0, 0), (-1, -1), pa_pad),
        ('TOPPADDING', (0, 0), (-1, -1), 2*mm),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 3*mm),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(bloc_pa4)

    # ── Bloc signature — coin droit, bas de page ──
    story.append(Spacer(1, 30*mm))

    sig_image_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "signature.png")
    sig_label = Paragraph(
        '<i>Expertise réalisée par :</i>',
        ParagraphStyle('SigLabel', fontName='Helvetica-Oblique', fontSize=8,
                       textColor=TEXT_LIGHT, alignment=TA_RIGHT, leading=10,
                       spaceAfter=2*mm))

    if os.path.exists(sig_image_path):
        sig_img = RLImage(sig_image_path, width=60*mm, height=14*mm)
        sig_block = Table([[sig_label], [sig_img]], colWidths=[65*mm])
        sig_block.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
    else:
        sig_name = Paragraph(
            "Fanny Lonqueu-Brochard",
            ParagraphStyle('SigName', fontName='Helvetica', fontSize=10,
                           textColor=TEXT_DARK, alignment=TA_RIGHT, leading=13))
        sig_block = Table([[sig_label], [sig_name]], colWidths=[65*mm])
        sig_block.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ]))

    # Wrapper pour aligner à droite
    sig_wrapper = Table([['', sig_block]], colWidths=[pa_content_w - 65*mm, 65*mm])
    sig_wrapper.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(sig_wrapper)

    # Note finale supprimée — le plan d'action clôt le rapport

    # Build
    doc.build(story, canvasmaker=NumberedCanvas)
    print(f"Rapport généré : {OUTPUT_PATH}")


if __name__ == "__main__":
    # ──────────────────────────────────────────
    # Usage :
    #   Mode démo :  python3 generate_report.py
    #   Mode prod :  python3 generate_report.py client.csv inventaire.xlsx
    # ──────────────────────────────────────────
    if len(sys.argv) == 3:
        csv_path = sys.argv[1]
        xlsx_path = sys.argv[2]

        print(f"Lecture CSV Tally : {csv_path}")
        tally_data = lire_csv_tally(csv_path)
        CLIENT_DATA.update(tally_data)

        print(f"Lecture Excel inventaire : {xlsx_path}")
        INVENTAIRE[:] = lire_inventaire_excel(xlsx_path)

        # Recalculer les données fixes après chargement
        CLIENT_DATA["date_rapport"] = datetime.now().strftime("%d/%m/%Y")
        CLIENT_DATA["ref_mission"] = generate_ref_mission(CLIENT_DATA["nom"])
        OUTPUT_PATH = generate_output_path()

        # Recalculer formats, apogées et totaux
        for b in INVENTAIRE:
            b["format"] = normaliser_format(b.get("format", "75cl"))
            if not b.get("apogee") or b["apogee"] == "":
                b["apogee"] = estimate_apogee(b)

        print(f"Client : {CLIENT_DATA['nom']} — Réf : {CLIENT_DATA['ref_mission']}")
        print(f"Inventaire : {len(INVENTAIRE)} références, {total_btls()} bouteilles")
    elif len(sys.argv) > 1:
        print("Usage : python3 generate_report.py [client.csv inventaire.xlsx]")
        sys.exit(1)
    else:
        print("Mode démo — données d'exemple")

    build_report()
