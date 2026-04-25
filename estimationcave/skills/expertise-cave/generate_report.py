#!/usr/bin/env python3
"""Générateur de rapport d'estimation de cave — version Jinja2 + Playwright.

Usage :
    python3 generate_report.py <inventaire.xlsx> <client.json> [output_dir]

- <inventaire.xlsx> : Excel avec feuille "Inventaire" (17 colonnes).
- <client.json>     : analyses expert rédigées par Fanny (calque sur
                      templates/sample_data.json). Fournit client,
                      perimetre, synthese, marche, plan_action,
                      accompagnement.
- [output_dir]      : défaut /mnt/user-data/outputs/.

La logique métier (lecture Excel, codes de justification, apogée,
référence mission) est conservée à l'identique. Tout le rendu ReportLab
est remplacé par un rendu HTML → PDF via le template Jinja2
templates/rapport.html et Playwright.
"""
from __future__ import annotations

import argparse
import asyncio
import json
import pathlib
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime

from jinja2 import Environment, FileSystemLoader, select_autoescape
from markupsafe import Markup
from playwright.async_api import async_playwright


HERE = pathlib.Path(__file__).parent
TEMPLATES_DIR = HERE / "templates"
TEMPLATE_NAME = "rapport.html"
DEFAULT_OUTPUT_DIR = pathlib.Path("/mnt/user-data/outputs")
CURRENT_YEAR = datetime.now().year

MOIS_FR = {
    1: "janvier", 2: "février", 3: "mars", 4: "avril", 5: "mai", 6: "juin",
    7: "juillet", 8: "août", 9: "septembre", 10: "octobre", 11: "novembre", 12: "décembre",
}

# ═════════════════════════════════════════════════════════════════
# Codes de justification (logique métier, identique à l'ancien ReportLab)
# ═════════════════════════════════════════════════════════════════
JUSTIFICATION_CODES = {
    # À CONSERVER
    1: "Fenêtre d'apogée non atteinte",
    2: "Cote en progression",
    3: "Référence rare ou confidentielle",
    4: "Millésime exceptionnel",
    5: "Grand format (meilleure conservation)",
    6: "Valeur de dégustation supérieure à la valeur de revente",
    # À VENDRE
    7: "Apogée atteinte ou dépassée",
    8: "Fenêtre de vente favorable",
    9: "État non compatible avec une conservation prolongée",
    10: "Référence accessible, ne se valorisera pas à l'avenir",
    11: "Millésime anniversaire (prime de liquidité)",
    12: "Format demi-bouteille (moins bonne conservation)",
    17: "Cote ayant atteint un plateau / tendance baissière",
    # À SURVEILLER
    13: "Maturité encore indéterminée",
    14: "Attractivité encore insuffisante sur le marché",
    15: "Millésime illisible ou inconnu — décision impossible sans vérification",
    16: "Signaux de marché contradictoires — tendance à confirmer",
}


# ═════════════════════════════════════════════════════════════════
# Lecture de l'Excel inventaire — logique métier reprise à l'identique
# ═════════════════════════════════════════════════════════════════
def lire_inventaire_excel(fichier_excel: pathlib.Path) -> list[dict]:
    """Lit la feuille 'Inventaire' avec alias de colonnes.

    Renvoie une liste de dicts avec les clés : bouteille, appellation, region,
    couleur, format, millesime (int), apogee, etat, cbo, qte, val_unit, reco,
    justification_code, duree_garde, canal_vente, reexamen, note_marche.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(fichier_excel), read_only=True)
    ws = wb["Inventaire"] if "Inventaire" in wb.sheetnames else wb.active
    print(f"  → Lecture de la feuille '{ws.title}' ({fichier_excel.name})")
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise RuntimeError("Excel vide")

    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    header_aliases = {
        "domaine": "Domaine", "domaine / cuvée": "Domaine", "domaine/cuvée": "Domaine",
        "domaine / cuvee": "Domaine", "nom": "Domaine", "cuvée": "Domaine",
        "appellation": "Appellation",
        "region": "Region", "région": "Region",
        "couleur": "Couleur", "type": "Couleur",
        "millesime": "Millesime", "millésime": "Millesime", "mill": "Millesime", "mill.": "Millesime",
        "format": "Format",
        "qte": "Qte", "qté": "Qte", "quantite": "Qte", "quantité": "Qte", "nb": "Qte",
        "etat": "Etat", "état": "Etat", "etat apparent": "Etat", "état apparent": "Etat",
        "cbo": "CBO", "caisse bois": "CBO",
        "val_unit": "Val_unit", "val unit": "Val_unit", "valeur unitaire": "Val_unit",
        "val. unit.": "Val_unit", "prix": "Val_unit", "valeur": "Val_unit",
        "reco": "Reco", "recommandation": "Reco", "reco.": "Reco",
        "justification_code": "Justification_code", "justification": "Justification_code",
        "code justification": "Justification_code", "code": "Justification_code",
        "duree_garde": "Duree_garde", "durée de garde": "Duree_garde", "duree garde": "Duree_garde",
        "durée garde": "Duree_garde", "garde": "Duree_garde",
        "apogee": "Apogee", "apogée": "Apogee",
        "note_marche": "Note_marche", "note marché": "Note_marche", "note marche": "Note_marche",
        "note de marché": "Note_marche",
        "canal_vente": "Canal_vente", "canal de vente": "Canal_vente", "canal vente": "Canal_vente",
        "canal": "Canal_vente",
        "reexamen": "Reexamen", "réexamen": "Reexamen", "rééxamen": "Reexamen",
        "date reexamen": "Reexamen", "date réexamen": "Reexamen",
    }
    headers = [header_aliases.get(h.lower().strip(), h) for h in raw_headers]

    # Normalisation Reco sans accent
    reco_map = {"A conserver": "À conserver", "A vendre": "À vendre", "A surveiller": "À surveiller"}

    def safe_str(val, default=""):
        if val is None:
            return default
        s = str(val).strip()
        return default if s.lower() in ("none", "") else s

    def safe_year(val, default=""):
        if val is None:
            return default
        try:
            return str(int(float(val)))
        except (ValueError, TypeError):
            return safe_str(val, default)

    inventaire = []
    for row in rows[1:]:
        if not row:
            continue
        padded = list(row) + [None] * (len(headers) - len(row))
        d = dict(zip(headers, padded))
        domaine_val = safe_str(d.get("Domaine"))
        if not domaine_val or domaine_val.lower() == "domaine" or not domaine_val[0].isalnum():
            continue

        mill_raw = d.get("Millesime")
        if mill_raw is None or str(mill_raw).strip().upper() in ("NM", "", "NONE"):
            mill = 0
        else:
            try:
                mill = int(float(mill_raw))
            except (ValueError, TypeError):
                mill = 0

        jcode_raw = d.get("Justification_code")
        try:
            jcode = int(float(jcode_raw)) if jcode_raw not in (None, "", "None") else None
        except (ValueError, TypeError):
            jcode = None

        reco_raw = safe_str(d.get("Reco"), "À surveiller")
        reco_val = reco_map.get(reco_raw, reco_raw)

        try:
            qte = int(float(d.get("Qte") or 1))
        except (ValueError, TypeError):
            qte = 1
        try:
            val_unit = int(float(d.get("Val_unit") or 0))
        except (ValueError, TypeError):
            val_unit = 0

        inventaire.append({
            "bouteille": domaine_val,
            "appellation": safe_str(d.get("Appellation")),
            "region": safe_str(d.get("Region")),
            "couleur": safe_str(d.get("Couleur"), "Rouge"),
            "format": safe_str(d.get("Format"), "75cl"),
            "millesime": mill,
            "apogee": safe_str(d.get("Apogee")),
            "etat": safe_str(d.get("Etat"), "Bon"),
            "cbo": safe_str(d.get("CBO"), "Non"),
            "qte": qte,
            "val_unit": val_unit,
            "reco": reco_val,
            "justification_code": jcode,
            "duree_garde": safe_str(d.get("Duree_garde")),
            "canal_vente": safe_str(d.get("Canal_vente")),
            "reexamen": safe_year(d.get("Reexamen")),
            "note_marche": safe_str(d.get("Note_marche")),
        })

    print(f"  → {len(inventaire)} références chargées")
    return inventaire


def generate_ref_mission(nom: str) -> str:
    """EST-AAAA-MM-XXX où XXX = 3 premières lettres du nom de famille."""
    cleaned = nom.replace("M.", "").replace("Mme", "").replace("Mme.", "").replace("et", "").strip()
    words = [w for w in cleaned.split() if len(w) > 1]
    last_word = words[-1] if words else "XXX"
    ascii_only = unicodedata.normalize("NFD", last_word).encode("ascii", "ignore").decode("ascii")
    letters = re.sub(r"[^A-Za-z]", "", ascii_only)
    code = letters[:3].upper() if len(letters) >= 3 else letters.upper().ljust(3, "X")
    now = datetime.now()
    return f"EST-{now.year}-{now.month:02d}-{code}"


def normaliser_format(fmt: str) -> str:
    """'1,5L' → '150cl', '0,375' → '37,5cl', '62' → '62cl'."""
    fmt = fmt.strip()
    if fmt.lower().endswith("cl"):
        return fmt
    has_l = "l" in fmt.lower()
    cleaned = fmt.lower().replace("l", "").replace(",", ".").strip()
    if not cleaned:
        return "75cl"
    try:
        val = float(cleaned)
        cl = val * 100 if (has_l or val < 10) else val
        return f"{int(cl)}cl" if cl == int(cl) else f"{cl:g}cl".replace(".", ",")
    except ValueError:
        return fmt


def estimate_apogee(bottle: dict) -> str:
    """Estime une fenêtre d'apogée si l'Excel n'en fournit pas.

    Renvoie 'AAAA–AAAA', 'À boire maintenant' ou 'À évaluer'. Reprise
    textuelle de la logique métier ReportLab.
    """
    mill = bottle.get("millesime", 0)
    if mill == 0:
        return "À boire maintenant"
    region = bottle.get("region", "").lower()
    appellation = bottle.get("appellation", "").lower()
    etat = bottle.get("etat", "Bon")

    if "grand cru" in appellation or "1er cru" in appellation or "premier cru" in appellation:
        window = (12, 35)
    elif region == "bourgogne":
        window = (8, 25)
    elif any(a in appellation for a in ("pauillac", "margaux", "saint-julien")):
        window = (12, 35)
    elif "sauternes" in appellation:
        window = (10, 50)
    elif region == "bordeaux":
        window = (8, 25)
    elif any(a in appellation for a in ("châteauneuf", "hermitage", "côte-rôtie")):
        window = (8, 25)
    elif region in ("vallée du rhône", "rhône"):
        window = (5, 18)
    elif region == "champagne":
        return "À boire maintenant"
    elif region == "alsace":
        window = (8, 25) if "grand cru" in appellation else (5, 15)
    elif region == "loire":
        window = (10, 40) if ("demi-sec" in bottle.get("bouteille", "").lower() or "moelleux" in appellation) else (5, 15)
    elif region in ("espagne", "italie"):
        keywords = ("ribera", "priorat", "barolo", "brunello", "toscana")
        window = (10, 30) if any(w in appellation for w in keywords) else (5, 18)
    elif region == "liban":
        window = (8, 30)
    else:
        return "À évaluer"

    if etat in ("Bon", "Moyen", "Abîmé", "À vérifier"):
        window = (window[0], max(window[0] + 3, window[1] - 5))

    start, end = mill + window[0], mill + window[1]
    if end < CURRENT_YEAR:
        return "À boire maintenant"
    return f"{start}–{end}"


# ═════════════════════════════════════════════════════════════════
# Formatage utilitaire pour le template
# ═════════════════════════════════════════════════════════════════
def fmt_int(n: int) -> str:
    """1234 → '1 234' (espace insécable entre milliers)."""
    return f"{int(n):,}".replace(",", "\u202f")


def _ceil10(x: float) -> int:
    """Plafond au multiple de 10 supérieur (4887,5 → 4890 ; 6325 → 6330)."""
    import math
    return int(math.ceil(x / 10)) * 10


def couleur_abbrev(c: str) -> str:
    """Version courte pour la cellule 'Coul.' du tableau inventaire."""
    c = c.strip()
    lower = c.lower()
    if lower.startswith("liquor"):
        return "Liquor."
    if lower.startswith("effervesc") or lower == "eff":
        return "Eff."
    return c


def etat_class(etat: str) -> str:
    """Helvetica-friendly class stem : 'Très bon' → 'tres-bon'."""
    ascii_ = unicodedata.normalize("NFD", etat).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z]+", "-", ascii_.lower()).strip("-")
    # On retombe sur les trois classes connues du design ; au-delà, on garde le slug.
    if slug in ("excellent", "tres-bon", "bon"):
        return slug
    return slug or "bon"


def orientation_class(reco: str) -> str:
    """Classe CSS : 'À conserver' → 'conserver', etc."""
    return {"À conserver": "conserver", "À vendre": "vendre", "À surveiller": "surveiller"}.get(reco, "surveiller")


def millesime_str(m: int) -> str:
    return str(m) if m else "NM"


# ═════════════════════════════════════════════════════════════════
# Calcul des agrégats à partir de l'inventaire
# ═════════════════════════════════════════════════════════════════
RECAP_ORDER = [
    ("À conserver", "",           "conserver"),
    ("À vendre",    "vendre",     "vendre"),
    ("À surveiller", "surveiller", "surveiller"),
]

# ─── Sous-régions : utilisé quand la cave est 100 % monorégion ──────────
# Pour une cave 100 % Bourgogne, afficher "Bordeaux 100 %" sur les deux
# graphiques région n'apporte rien. On bascule sur un découpage par
# sous-région (Côte de Nuits vs Côte de Beaune, etc.). Les mots-clés
# cherchent dans l'appellation, lowercasée et sans accents sensibles.
#
# Règle canonique : une région = N sous-régions ESSENTIELLES. Tout ce qui
# n'y tombe pas (appellations régionales, Hautes Côtes, grandes appellations
# inter-communales non listées…) atterrit dans "{Région} (autre)".
# Les listes sont explicites et exhaustives — on ne liste QUE les
# sous-régions qu'on veut voir apparaître comme catégorie à part entière.
#
# Attention à l'ordre de matching dans `subregion_of()` :
# - "nuits-saint-georges" doit matcher AVANT toute variante générique "nuits".
# - Les appellations inter-communales (Chambertin, Musigny, Bâtard-Montrachet)
#   doivent matcher avant les noms de commune courts.
#
SUBREGION_KEYWORDS: dict[str, list[tuple[str, list[str]]]] = {
    "Bourgogne": [
        ("Côte de Nuits", [
            # Communes
            "gevrey-chambertin", "morey-saint-denis", "chambolle-musigny",
            "vosne-romanée", "vosne romanée", "flagey-échézeaux",
            "nuits-saint-georges", "nuits saint-georges",
            "marsannay", "fixin", "vougeot",
            # Grands Crus inter-communaux (Côte de Nuits)
            "échézeaux", "echezeaux", "chambertin", "charmes-chambertin",
            "mazis-chambertin", "mazoyères-chambertin", "griotte-chambertin",
            "latricières-chambertin", "ruchottes-chambertin", "chapelle-chambertin",
            "clos de bèze", "clos de tart", "clos de la roche",
            "clos saint-denis", "clos des lambrays", "bonnes mares",
            "clos de vougeot", "clos-de-vougeot", "musigny", "romanée-conti",
            "richebourg", "la tâche", "grands-échézeaux",
        ]),
        ("Côte de Beaune", [
            # Communes
            "ladoix", "aloxe-corton", "pernand-vergelesses",
            "savigny-lès-beaune", "savigny les beaune",
            "chorey-lès-beaune", "chorey les beaune",
            "beaune", "pommard", "volnay", "monthelie", "auxey-duresses",
            "saint-romain", "meursault",
            "puligny-montrachet", "puligny montrachet",
            "chassagne-montrachet", "chassagne montrachet",
            "saint-aubin", "saint aubin", "santenay", "maranges",
            # Grands Crus inter-communaux (Côte de Beaune)
            "corton", "charlemagne", "corton-charlemagne",
            "montrachet", "bâtard-montrachet", "batard-montrachet",
            "bienvenues-bâtard-montrachet", "criots-bâtard-montrachet",
            "chevalier-montrachet",
        ]),
        ("Côte Chalonnaise", [
            "bouzeron", "rully", "mercurey", "givry", "montagny",
        ]),
        ("Mâconnais", [
            "mâcon", "macon", "pouilly-fuissé", "pouilly fuisse",
            "pouilly-vinzelles", "pouilly-loché",
            "saint-véran", "saint veran", "viré-clessé", "vire-clesse",
        ]),
        ("Chablisien", [
            "chablis", "petit chablis",
        ]),
    ],
    "Bordeaux": [
        ("Médoc", [
            "pauillac", "saint-julien", "saint julien",
            "saint-estèphe", "saint estephe", "saint-estephe",
            "margaux", "moulis", "listrac",
            "médoc", "medoc", "haut-médoc", "haut-medoc",
        ]),
        ("Pessac-Léognan / Graves", [
            "pessac-léognan", "pessac leognan", "pessac-leognan",
            "graves",
        ]),
        ("Libournais (Saint-Émilion / Pomerol)", [
            "saint-émilion", "saint emilion", "saint-emilion",
            "pomerol", "lalande-de-pomerol",
            "fronsac", "canon-fronsac",
        ]),
        ("Sauternais", [
            "sauternes", "barsac", "cérons", "sainte-croix-du-mont",
            "loupiac",
        ]),
    ],
    "Vallée du Rhône": [
        ("Rhône septentrional", [
            "côte-rôtie", "cote-rotie", "côte rôtie",
            "condrieu", "saint-joseph", "crozes-hermitage", "hermitage",
            "cornas", "saint-péray", "saint peray",
            "château-grillet", "chateau-grillet",
        ]),
        ("Rhône méridional", [
            "châteauneuf-du-pape", "chateauneuf-du-pape", "châteauneuf",
            "gigondas", "vacqueyras", "lirac", "tavel", "rasteau",
            "beaumes-de-venise",
        ]),
    ],
    "Champagne": [
        ("Montagne de Reims", [
            "ambonnay", "bouzy", "verzenay", "verzy", "mailly", "aÿ",
        ]),
        ("Côte des Blancs", [
            "cramant", "avize", "oger", "le mesnil", "chouilly",
        ]),
        ("Vallée de la Marne", [
            "hautvillers", "damery", "cumières",
        ]),
    ],
    "Loire": [
        ("Pays Nantais", ["muscadet"]),
        ("Anjou-Saumur", [
            "anjou", "saumur", "savennières",
            "bonnezeaux", "quarts-de-chaume", "coteaux du layon", "coteaux-du-layon",
        ]),
        ("Touraine", [
            "vouvray", "montlouis", "chinon", "bourgueil",
            "saint-nicolas-de-bourgueil", "touraine",
        ]),
        ("Centre-Loire", [
            "sancerre", "pouilly-fumé", "pouilly fume",
            "menetou-salon", "reuilly", "quincy",
        ]),
    ],
}


def subregion_of(region: str, appellation: str) -> str | None:
    """Renvoie la sous-région pour un couple (région, appellation) si connue."""
    table = SUBREGION_KEYWORDS.get(region)
    if not table:
        return None
    app_lower = (appellation or "").lower()
    for subregion, keywords in table:
        if any(k in app_lower for k in keywords):
            return subregion
    return None

COULEUR_COLORS = [
    # Couleur libellée, couleur CSS, texte, small ?, label legend
    ("Rouge",         "#6A1F2E",              "#FAF6F0",         False, "Rouge"),
    ("Blanc",         "#C5A258",              "var(--bordeaux)", False, "Blanc sec"),
    ("Rosé",          "#E8A0B0",              "var(--bordeaux)", True,  "Rosé"),
    ("Liquoreux",     "#8a6a2d",              "var(--fond)",     True,  "Liquoreux"),
    ("Effervescent",  "rgba(138,106,45,0.6)", "var(--fond)",     True,  "Effervescent"),
]

GARDE_GROUPS = [
    ('À boire&nbsp;<span class="garde-sub">— consommation immédiate</span>',                      "#8B3A3A"),
    ('Proche apogée (2/4 ans)&nbsp;<span class="garde-sub">— fenêtre qui s\'ouvre bientôt</span>', "#B8883A"),
    ('En développement (5/10 ans)&nbsp;<span class="garde-sub">— patience requise</span>',         "#5A7A4B"),
    ('Garde longue (10 ans et plus)&nbsp;<span class="garde-sub">— patrimoine à long terme</span>', "#2D1B2E"),
]


def apogee_bucket(apogee: str) -> int:
    """Retourne l'index (0-3) dans GARDE_GROUPS selon la fenêtre d'apogée."""
    apogee = (apogee or "").strip()
    if apogee.lower().startswith(("à boire", "a boire")):
        return 0
    m = re.match(r"(\d{4})\s*[–\-]\s*(\d{4})", apogee)
    if not m:
        return 0
    start, end = int(m.group(1)), int(m.group(2))
    mid = (start + end) / 2
    diff = mid - CURRENT_YEAR
    if diff <= 0:
        return 0
    if diff <= 3:
        return 1
    if diff <= 10:
        return 2
    return 3


def _is_monoregion(inv: list[dict]) -> tuple[bool, str | None]:
    """Renvoie (True, 'Bourgogne') si toutes les lignes partagent la même
    région, (False, None) sinon."""
    regions = {b["region"] for b in inv if b["region"]}
    if len(regions) == 1:
        return True, next(iter(regions))
    return False, None


def _group_by_geo(inv: list[dict], mono_region: str | None) -> tuple[dict, dict, str]:
    """Renvoie (by_geo_vol, by_geo_val, axis_label).

    En monorégion, on regroupe par sous-région ("Côte de Nuits", "Côte de
    Beaune"…). Sinon, par région. axis_label sert à titrer les graphiques
    ("région" ou "sous-région").
    """
    by_vol: dict[str, int] = defaultdict(int)
    by_val: dict[str, int] = defaultdict(int)
    if mono_region:
        fallback = f"{mono_region} (autre)"
        for b in inv:
            sub = subregion_of(mono_region, b["appellation"]) or fallback
            by_vol[sub] += b["qte"]
            by_val[sub] += b["qte"] * b["val_unit"]
        return by_vol, by_val, "sous-région"
    for b in inv:
        r = b["region"] or "Autres"
        by_vol[r] += b["qte"]
        by_val[r] += b["qte"] * b["val_unit"]
    return by_vol, by_val, "région"


def build_synthese(inv: list[dict]) -> dict:
    total_val = sum(b["qte"] * b["val_unit"] for b in inv)
    total_btl = sum(b["qte"] for b in inv)
    mono, region_name = _is_monoregion(inv)

    # Par région (ou sous-région si monorégion) — barres page 3
    _, by_geo_val, axis = _group_by_geo(inv, region_name if mono else None)
    regions_sorted = sorted(by_geo_val.items(), key=lambda x: -x[1])
    max_val = regions_sorted[0][1] if regions_sorted else 1
    regions_valeur = [
        {
            "nom": nom,
            "pct_fmt": f"{val / total_val * 100:.1f}".replace(".", ",") if total_val else "0,0",
            "width": round(val / max_val * 100, 1) if max_val else 0,
        }
        for nom, val in regions_sorted
    ]
    geo_axis = axis  # "région" ou "sous-région"

    # Répartition par couleur : volume et valeur
    by_col_vol = defaultdict(int)
    by_col_val = defaultdict(int)
    for b in inv:
        key = b["couleur"].strip() or "Rouge"
        by_col_vol[key] += b["qte"]
        by_col_val[key] += b["qte"] * b["val_unit"]

    def color_segments(buckets: dict, total: int) -> list[dict]:
        segments = []
        for label, color, text_color, small, _ in COULEUR_COLORS:
            if label not in buckets or buckets[label] == 0:
                continue
            pct = buckets[label] / total * 100 if total else 0
            segments.append({
                "pct_fmt": f"{pct:.0f}",
                "width": round(pct, 1),
                "color": color,
                "text_color": text_color,
                "small": small or pct < 10,
            })
        return segments

    # Légende dynamique : on n'affiche que les couleurs effectivement
    # présentes dans la cave (sinon Rosé / Liquoreux apparaissent à tort).
    color_legend = [
        {"label": legend, "color": color}
        for label, color, _tc, _s, legend in COULEUR_COLORS
        if by_col_vol.get(label, 0) > 0 or by_col_val.get(label, 0) > 0
    ]

    color_split = {
        "volume":  color_segments(by_col_vol, total_btl),
        "valeur":  color_segments(by_col_val, total_val),
        "legende": color_legend,
    }

    # Recap orientations : lots (≠ références pour les 3 catégories)
    recap = []
    for label, dot_class, _ in RECAP_ORDER:
        lots = [b for b in inv if b["reco"] == label]
        btl = sum(b["qte"] for b in lots)
        val = sum(b["qte"] * b["val_unit"] for b in lots)
        recap.append({
            "label": label,
            "dot_class": dot_class,
            "lots": len(lots),
            "btl": btl,
            "valeur_fmt": fmt_int(val),
        })

    return {
        "valeur_totale": total_val,
        "valeur_totale_fmt": fmt_int(total_val),
        # Fourchette asymétrique -15 % / +10 % — reprise du comportement ReportLab
        # historique : la borne basse reflète les décotes d'adjudication, la haute
        # est volontairement prudente car les pics spéculatifs sont rares.
        # Les deux bornes sont arrondies au multiple de 10 supérieur pour une
        # lecture "ronde" cohérente avec l'ancien rapport.
        "fourchette_min_fmt": fmt_int(_ceil10(total_val * 0.85)),
        "fourchette_max_fmt": fmt_int(_ceil10(total_val * 1.10)),
        "geo_axis": geo_axis,
        "regions_valeur": regions_valeur,
        "color_split": color_split,
        "recap": recap,
    }


def build_repartition(inv: list[dict], synthese: dict) -> dict:
    total_val = synthese["valeur_totale"]
    total_btl = sum(b["qte"] for b in inv)
    mono, region_name = _is_monoregion(inv)

    # Volume / valeur par région (ou par sous-région si la cave est monorégion)
    by_vol, by_val, axis_label = _group_by_geo(inv, region_name if mono else None)

    def to_list(buckets: dict, total: int) -> list[dict]:
        sorted_items = sorted(buckets.items(), key=lambda x: -x[1])
        return [
            {
                "label": nom,
                "pct": round(v / total * 100, 1) if total else 0,
                "pct_fmt": f"{v / total * 100:.1f}".replace(".", ",") if total else "0,0",
                "value_fmt": fmt_int(v),
            }
            for nom, v in sorted_items
        ]

    volume_region = to_list(by_vol, total_btl)
    valeur_region = to_list(by_val, total_val)

    # Potentiel de garde — segments + légende
    buckets = [0, 0, 0, 0]
    for b in inv:
        buckets[apogee_bucket(b["apogee"])] += b["qte"]

    segments = []
    legende = []
    legende_details = [
        "fenêtre de consommation ouverte",
        "2 à 4 ans avant le pic",
        "encore en phase de garde active",
        "potentiel > 10 ans",
    ]
    plain_labels = ["À boire", "Proche apogée", "En développement", "Garde longue"]
    for idx, ((_rich_label, color), count) in enumerate(zip(GARDE_GROUPS, buckets)):
        if count == 0:
            continue
        pct = count / total_btl * 100 if total_btl else 0
        segments.append({"label": plain_labels[idx], "pct": round(pct, 1), "color": color})
        legende.append({
            "label": plain_labels[idx],
            "detail": f"{count} btl · {legende_details[idx]}",
            "pct_fmt": f"{pct:.0f}",
            "color": color,
        })

    # Reco split (page 7)
    reco_colors = {"À conserver": "#5A7A4B", "À surveiller": "#B8883A", "À vendre": "#8B3A3A"}
    reco_order = ["À conserver", "À surveiller", "À vendre"]
    vol_seg, val_seg = [], []
    vol_by_reco, val_by_reco, btl_by_reco = defaultdict(int), defaultdict(int), defaultdict(int)
    for b in inv:
        vol_by_reco[b["reco"]] += b["qte"]
        val_by_reco[b["reco"]] += b["qte"] * b["val_unit"]
        btl_by_reco[b["reco"]] += b["qte"]

    def reco_segment(buckets: dict, total: int) -> list[dict]:
        out = []
        for label in reco_order:
            val = buckets.get(label, 0)
            if val == 0:
                continue
            pct = val / total * 100 if total else 0
            out.append({
                "label": label.replace("À ", "").capitalize(),
                "pct": round(pct, 1),
                "pct_int": round(pct),
                "color": reco_colors[label],
            })
        return out

    reco_legende = []
    for label in reco_order:
        btl = btl_by_reco.get(label, 0)
        val = val_by_reco.get(label, 0)
        if btl == 0 and val == 0:
            continue
        detail_suffix = {
            "À conserver": "cœur patrimonial, à maintenir en cave.",
            "À surveiller": "à réexaminer selon les échéances indiquées.",
            "À vendre": "fenêtre de cession favorable.",
        }[label]
        reco_legende.append({
            "label": label.replace("À ", "").capitalize(),
            "detail": f"{btl} btl / {fmt_int(val)} € — {detail_suffix}",
            "color": reco_colors[label],
        })

    reco_split = {
        "volume": reco_segment(vol_by_reco, total_btl),
        "valeur": reco_segment(val_by_reco, total_val),
        "valeur_total_fmt": fmt_int(total_val),
        "legende": reco_legende,
    }

    # Top 5 unitaire et totale
    by_unit = sorted(inv, key=lambda b: -b["val_unit"])[:5]
    by_total = sorted(inv, key=lambda b: -(b["qte"] * b["val_unit"]))[:5]
    top_unitaire = [
        {
            "nom": f"{b['bouteille']} {millesime_str(b['millesime'])}".strip(),
            "detail": b["appellation"] or "",
            "value_fmt": fmt_int(b["val_unit"]),
        }
        for b in by_unit
    ]
    top_totale = [
        {
            "nom": f"{b['bouteille']} {millesime_str(b['millesime'])}".strip(),
            "detail": f"{b['qte']} btl",
            "value_fmt": fmt_int(b["qte"] * b["val_unit"]),
        }
        for b in by_total
    ]
    top_concentration = sum(b["qte"] * b["val_unit"] for b in by_total)
    top_concentration_pct = round(top_concentration / total_val * 100) if total_val else 0

    return {
        "volume_region": volume_region,
        "valeur_region_base_fmt": fmt_int(total_val),
        "valeur_region": valeur_region,
        "geo_axis": axis_label,
        "mono_region": region_name if mono else None,
        "garde": {"segments": segments, "legende": legende},
        "reco_split": reco_split,
        "top_unitaire": top_unitaire,
        "top_totale": top_totale,
        "top_concentration_pct": top_concentration_pct,
    }


def build_inventaire_rows(inv: list[dict]) -> list[dict]:
    """Transforme l'inventaire brut en lignes prêtes pour le template p4.

    Tri par millésime croissant (NM = 0 en tête), puis par domaine
    alphabétique pour les ex æquo : permet de lire la cave dans l'ordre
    chronologique et facilite les comparaisons inter-millésimes.
    """
    rows = []
    # Tri par millésime ASC, NM (millesime=0) renvoyé en fin de liste,
    # puis par domaine alphabétique pour les ex æquo.
    inv = sorted(inv, key=lambda b: (b.get("millesime") or 9999, b.get("bouteille", "")))
    for b in inv:
        rows.append({
            "domaine": b["bouteille"],
            "cuvee": None,  # on laisse l'appellation dans sa propre colonne
            "appellation": b["appellation"] or "—",
            "region": b["region"] or "—",
            "couleur": couleur_abbrev(b["couleur"]),
            "millesime": millesime_str(b["millesime"]),
            "etat": b["etat"] or "Bon",
            "etat_class": etat_class(b["etat"] or "Bon"),
            "qte": b["qte"],
            "unit_fmt": fmt_int(b["val_unit"]),
            "total_fmt": fmt_int(b["qte"] * b["val_unit"]),
            "orientation": b["reco"],
            "orientation_class": orientation_class(b["reco"]),
        })
    return rows


def build_recommandations(inv: list[dict]) -> list[dict]:
    """Groupes conserver / vendre / surveiller avec cartes détaillées."""
    groups = []
    for label, class_mod, _orient_class in RECAP_ORDER:
        lots = [b for b in inv if b["reco"] == label]
        # Tri par millésime ASC, NM en fin, puis par domaine alphabétique.
        # Cohérent avec l'inventaire détaillé p4.
        lots.sort(key=lambda b: (b.get("millesime") or 9999, b.get("bouteille", "")))
        if not lots:
            continue
        nb_btl = sum(b["qte"] for b in lots)
        val = sum(b["qte"] * b["val_unit"] for b in lots)

        lot_blocks = []
        for b in lots:
            code = b.get("justification_code")
            code_title = JUSTIFICATION_CODES.get(code) if code else None
            note = b.get("note_marche") or ""
            if code_title and note:
                justification = f"<b>{code_title}.</b> {note}"
            elif code_title:
                justification = f"<b>{code_title}.</b>"
            else:
                justification = note or "[Justification à compléter]"

            mill = millesime_str(b["millesime"])
            appellation_parts = [b["appellation"] or "—", mill, b.get("format") or "75cl"]
            if b.get("cbo", "").lower() == "oui":
                appellation_parts.append("CBO")
            appellation = " · ".join(appellation_parts)

            meta = []
            if b.get("apogee"):
                meta.append({"lbl": "Apogée", "val": b["apogee"]})
            if label == "À vendre" and b.get("canal_vente"):
                meta.append({"lbl": "Canal de vente", "val": b["canal_vente"]})
            if label == "À surveiller" and b.get("reexamen"):
                meta.append({"lbl": "Réexamen", "val": b["reexamen"]})

            lot_blocks.append({
                "nom": b["bouteille"],
                "appellation": appellation,
                "qte": b["qte"],
                "valeur_fmt": fmt_int(b["qte"] * b["val_unit"]),
                "garde": b.get("duree_garde") if label == "À conserver" else None,
                "justification": justification,
                "meta": meta,
            })

        groups.append({
            "titre": label,
            "class_mod": class_mod,
            "nb_bouteilles": nb_btl,
            "valeur_fmt": fmt_int(val),
            "lots": lot_blocks,
        })
    return groups


def build_potentiel_garde(inv: list[dict]) -> dict:
    """Les 4 fenêtres temporelles avec lignes détaillées."""
    buckets: list[list[dict]] = [[], [], [], []]
    for b in inv:
        buckets[apogee_bucket(b["apogee"])].append(b)

    groupes = []
    for idx, (titre, color) in enumerate(GARDE_GROUPS):
        bottles = buckets[idx]
        if not bottles:
            continue
        bottles.sort(key=lambda b: (b["millesime"] or 9999, b["bouteille"]))
        nb_btl = sum(b["qte"] for b in bottles)
        val = sum(b["qte"] * b["val_unit"] for b in bottles)
        lignes = [
            {
                "domaine": b["bouteille"],
                "appellation": b["appellation"] or "—",
                "millesime": millesime_str(b["millesime"]),
                "qte": b["qte"],
                "apogee": b["apogee"] or "—",
                "orientation": b["reco"],
                "orientation_class": orientation_class(b["reco"]),
            }
            for b in bottles
        ]
        groupes.append({
            "titre": titre,
            "color": color,
            "nb_bouteilles": nb_btl,
            "valeur_fmt": fmt_int(val),
            "lignes": lignes,
        })

    return {
        "intro": ("Pour chaque lot, l'apogée estimée indique la fenêtre optimale de consommation. "
                  "Ces estimations sont établies à titre indicatif sur la base du millésime, de "
                  "l'appellation et du style du domaine."),
        "groupes": groupes,
    }


def build_liquidite_globale(inv: list[dict]) -> str:
    """Approximation : Élevée si ≥ 50 % de la valeur sur régions très liquides."""
    liquide = {"bourgogne", "bordeaux", "champagne"}
    total = sum(b["qte"] * b["val_unit"] for b in inv) or 1
    liq_val = sum(b["qte"] * b["val_unit"] for b in inv if b["region"].lower() in liquide)
    pct = liq_val / total
    if pct >= 0.6:
        return "Élevée"
    if pct >= 0.3:
        return "Moyenne"
    return "Faible"


# ═════════════════════════════════════════════════════════════════
# Rendu (Jinja2 + Playwright)
# ═════════════════════════════════════════════════════════════════
def dropcap(html: str) -> Markup:
    """Première lettre visible wrappée dans <span class='dropcap'>."""
    html = str(html)
    m = re.search(r"^((?:\s|<[^>]+>)*)(\w)", html)
    if not m:
        return Markup(html)
    prefix, letter = m.group(1), m.group(2)
    return Markup(f'{prefix}<span class="dropcap">{letter}</span>{html[len(prefix) + 1:]}')


def make_env() -> Environment:
    env = Environment(
        loader=FileSystemLoader(str(TEMPLATES_DIR)),
        autoescape=select_autoescape(["html"]),
    )
    env.filters["dropcap"] = dropcap
    return env


def split_paragraphs(text: str) -> list[str]:
    """Sépare un texte Python multi-ligne en paragraphes sur \\n\\n."""
    return [p.strip() for p in (text or "").split("\n\n") if p.strip()]


def build_render_context(inv: list[dict], client_json: dict) -> dict:
    """Assemble le dict complet passé au template Jinja."""
    synthese = build_synthese(inv)
    repartition = build_repartition(inv, synthese)
    inventaire_rows = build_inventaire_rows(inv)
    recommandations = build_recommandations(inv)
    potentiel = build_potentiel_garde(inv)

    liquidite = (client_json.get("synthese") or {}).get("liquidite_globale") or build_liquidite_globale(inv)

    client = dict(client_json.get("client", {}))
    client.setdefault("ref_dossier", generate_ref_mission(client.get("nom", "Client")))

    perimetre = dict(client_json.get("perimetre", {}))
    perimetre["nb_references"] = len(inv)
    perimetre["nb_bouteilles"] = sum(b["qte"] for b in inv)
    perimetre.setdefault("base_estimation", "Estimation basée sur photos")
    perimetre.setdefault("conditions", "Sur éléments fournis par le client")
    perimetre.setdefault("conservation", "Non communiquée")
    perimetre.setdefault("provenance", "Non communiquée")

    rapport = dict(client_json.get("rapport", {}))
    now = datetime.now()
    rapport.setdefault("date_emission", f"{now.day} {MOIS_FR[now.month]} {now.year}")
    rapport.setdefault("validite", "Estimation valable 30 jours")

    expert = client_json.get("expert") or {
        "nom": "Fanny Lonqueu-Brochard",
        "titre": "Expert indépendante · estimationcave.com",
    }

    synthese["liquidite_globale"] = liquidite

    return {
        "client": client,
        "expert": expert,
        "rapport": rapport,
        "perimetre": perimetre,
        "synthese": synthese,
        "inventaire": inventaire_rows,
        "marche": client_json.get("marche", default_marche()),
        "repartition": repartition,
        "recommandations": recommandations,
        "potentiel_garde": potentiel,
        "plan_action": client_json.get("plan_action", []),
        "accompagnement": client_json.get("accompagnement", default_accompagnement()),
    }


def default_marche() -> dict:
    return {"lede": "", "blocs": []}


def default_accompagnement() -> dict:
    return {
        "missions": {
            "titre": "Missions complémentaires disponibles",
            "tag": "Sur mesure",
            "intro": ("Ce rapport peut être complété par des missions d'accompagnement sur mesure, "
                      "selon vos besoins et l'évolution de votre projet."),
            "liste": [
                {"titre": "Inspection physique des flacons",
                 "description": "Examen approfondi en cave, avec déplacement ou recours à un mandataire local. Permet de confirmer l'état réel des bouteilles et d'affiner l'estimation."},
                {"titre": "Accompagnement à la mise en vente",
                 "description": "Sélection de la maison de vente ou plateforme adaptée, constitution du dossier de présentation des lots, coordination avec les experts."},
                {"titre": "Suivi des adjudications",
                 "description": "Présence ou suivi à distance lors de la vente, compte-rendu détaillé des résultats et analyse des écarts avec l'estimation initiale."},
                {"titre": "Réévaluation périodique",
                 "description": "Mise à jour annuelle ou semestrielle de l'estimation, intégrant les évolutions du marché et l'état de conservation de la cave."},
                {"titre": "Conseil à l'achat",
                 "description": "Identification de références à acquérir pour compléter, rééquilibrer ou valoriser la cave, selon un budget et des objectifs définis."},
            ],
            "devis": "Ces missions font l'objet d'un devis personnalisé. Pour toute demande : <b>contact@estimationcave.com</b>",
        },
        "confidentialite": {
            "titre": "Confidentialité réciproque",
            "tag": "Cadre contractuel",
            "corps": ("Ce rapport est établi dans le cadre d'une relation de stricte confidentialité. "
                      "Le client s'engage à ne pas diffuser ce document à des tiers non autorisés. "
                      "L'expert s'engage en retour à ne divulguer aucune information relative à "
                      "l'existence, la composition, la localisation ou la valeur de la cave à quelque "
                      "tiers que ce soit, sans accord écrit préalable du client."),
        },
        "validite": {
            "titre": "Validité et contact",
            "tag": "Clôture",
            "corps": ("Les estimations contenues dans ce rapport sont valables <b>30 jours</b> à "
                      "compter de la date d'émission. Au-delà, une réévaluation peut être demandée. "
                      "Pour toute question : <b>contact@estimationcave.com</b>"),
        },
    }


async def render_pdf(context: dict, output_path: pathlib.Path) -> None:
    """Rendu mono-passe : la pagination "X / Y" vient désormais de
    @page @bottom-right { content: counter(page) " / " counter(pages) }
    — le CSS Paged Media de Chromium compte automatiquement les A4
    physiques, sans qu'on ait besoin d'un pré-rendu."""
    env = make_env()
    html = env.get_template(TEMPLATE_NAME).render(**context)
    scratch = TEMPLATES_DIR / "_rendered.html"
    scratch.write_text(html)
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.goto(scratch.as_uri(), wait_until="networkidle")
            await page.pdf(
                path=str(output_path),
                format="A4",
                print_background=True,
                margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
            )
            await browser.close()
    finally:
        scratch.unlink(missing_ok=True)
    n_pages = len(re.findall(rb"/Type\s*/Page[^s]", output_path.read_bytes()))
    print(f"  → PDF : {output_path} ({output_path.stat().st_size:,} octets, {n_pages} pages)")


# ═════════════════════════════════════════════════════════════════
# Nommage du fichier de sortie
# ═════════════════════════════════════════════════════════════════
def build_output_path(nom: str, ref: str, output_dir: pathlib.Path) -> pathlib.Path:
    cleaned = nom.replace("M.", "").replace("Mme", "").replace("et", "").strip()
    words = [w for w in cleaned.split() if len(w) > 1]
    nom_famille = unicodedata.normalize("NFD", words[-1] if words else "CLIENT")
    nom_ascii = nom_famille.encode("ascii", "ignore").decode("ascii").upper()
    now = datetime.now()
    mois_annee = f"{MOIS_FR[now.month].capitalize()}_{now.year}"
    filename = f"Audit_et_estimation_de_cave__{nom_ascii}__{ref}__{mois_annee}.pdf"
    return output_dir / filename


# ═════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════
def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("inventaire", type=pathlib.Path, help="Excel inventaire (feuille 'Inventaire').")
    parser.add_argument("client_json", type=pathlib.Path, help="JSON des analyses expert.")
    parser.add_argument("output_dir", type=pathlib.Path, nargs="?", default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    print("→ Lecture de l'inventaire Excel")
    inv = lire_inventaire_excel(args.inventaire)
    for b in inv:
        b["format"] = normaliser_format(b["format"])
        if not b["apogee"]:
            b["apogee"] = estimate_apogee(b)

    print("→ Lecture des analyses expert")
    client_json = json.loads(args.client_json.read_text())

    print("→ Calcul des agrégats")
    context = build_render_context(inv, client_json)
    ref = context["client"]["ref_dossier"]
    nom = context["client"]["nom"]

    args.output_dir.mkdir(parents=True, exist_ok=True)
    out = build_output_path(nom, ref, args.output_dir)

    print(f"→ Rendu PDF ({nom}, {ref})")
    asyncio.run(render_pdf(context, out))
    print("→ Terminé.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
